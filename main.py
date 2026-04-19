from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

MACRS_5YR = [0.20, 0.32, 0.192, 0.1152, 0.1152, 0.0576]

def _irr_robust(cfs, guess=0.08):
    """여러 초기값으로 Newton 반복 → 양수 수렴값 반환"""
    import numpy as np
    def newton(g):
        r = g
        for _ in range(2000):
            npv  = sum(cf/(1+r)**t for t,cf in enumerate(cfs))
            dnpv = sum(-t*cf/(1+r)**(t+1) for t,cf in enumerate(cfs))
            if abs(dnpv) < 1e-12: break
            r_new = r - npv/dnpv
            if r_new <= -0.999: r_new = 0.001
            if abs(r_new - r) < 1e-8: return r_new
            r = r_new
        return r
    for g in [guess, 0.01, 0.03, 0.05, 0.10, 0.15]:
        r = newton(g)
        if r > 0:
            chk = sum(cf/(1+r)**t for t,cf in enumerate(cfs))
            if abs(chk) < 500:  # $500K 오차 허용
                return r
    try:
        r0 = float(npf.irr(cfs))
        return r0 if not np.isnan(r0) else 0.0
    except Exception:
        return 0.0
import requests
import jwt
import os
import datetime
import json
import tempfile
from pyxlsb import open_workbook

app = FastAPI(title="HWR Dashboard API")

# ── CORS (대시보드에서 호출 허용) ──────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 추후 shinjo99.github.io 로 제한
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── 설정 ──────────────────────────────────────────
JWT_SECRET     = os.environ.get("JWT_SECRET", "hwr-secret-change-this")
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
FB_URL      = os.environ.get("FB_URL", "https://team-dashboard-c0d7b-default-rtdb.asia-southeast1.firebasedatabase.app")
FB_SECRET   = os.environ.get("FB_SECRET", "")  # Firebase Database Secret
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")  # https://fred.stlouisfed.org/docs/api/api_key.html

# ── 사용자 계정 (환경변수로 관리) ─────────────────
# 형식: "email:password:role,email:password:role"
# 예: "team@hwr.com:hanwha2024:viewer,admin@hwr.com:admin1234:admin"
def get_users():
    raw = os.environ.get("USERS", "team@hwr.com:hanwha2024:viewer")
    users = {}
    for entry in raw.split(","):
        parts = entry.strip().split(":")
        if len(parts) >= 3:
            email, password, role = parts[0], parts[1], parts[2]
            users[email] = {"password": password, "role": role}
    return users

# ── JWT 헬퍼 ──────────────────────────────────────
def create_token(email: str, role: str) -> str:
    payload = {
        "email": email,
        "role": role,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def verify_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="토큰이 만료되었습니다. 다시 로그인하세요.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="유효하지 않은 토큰입니다.")

security = HTTPBearer()

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    return verify_token(credentials.credentials)

def require_admin(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    return user

# ── Firebase 헬퍼 ──────────────────────────────────
def fb_auth_param():
    if FB_SECRET:
        return {"auth": FB_SECRET}
    return {}

def fb_read(path: str):
    try:
        res = requests.get(
            f"{FB_URL}/{path}.json",
            params=fb_auth_param(),
            timeout=5
        )
        return res.json() or {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 읽기 오류: {str(e)}")

def fb_write(path: str, data: dict):
    try:
        res = requests.patch(
            f"{FB_URL}/{path}.json",
            json=data,
            params=fb_auth_param(),
            timeout=5
        )
        if res.status_code != 200:
            raise HTTPException(status_code=500, detail="DB 저장 실패")
        return res.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 쓰기 오류: {str(e)}")

def fb_put(path: str, data: dict):
    try:
        res = requests.put(
            f"{FB_URL}/{path}.json",
            json=data,
            params=fb_auth_param(),
            timeout=5
        )
        if res.status_code != 200:
            raise HTTPException(status_code=500, detail="DB 저장 실패")
        return res.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 쓰기 오류: {str(e)}")

def fb_patch(path: str, data: dict):
    """Firebase PATCH — 필드 일부만 업데이트"""
    try:
        requests.patch(
            f"{FB_URL}/{path}.json",
            json=data,
            params=fb_auth_param(),
            timeout=5
        )
    except Exception:
        pass

# ══════════════════════════════════════════════════
#  인증
# ══════════════════════════════════════════════════
class LoginRequest(BaseModel):
    email: str
    password: str

class ValuationCalcRequest(BaseModel):
    project_id: str = ""
    inputs: dict = {}

@app.post("/auth/login")
def login(req: LoginRequest):
    users = get_users()
    user = users.get(req.email)
    if not user or user["password"] != req.password:
        raise HTTPException(status_code=401, detail="이메일 또는 비밀번호가 틀렸습니다.")
    token = create_token(req.email, user["role"])
    return {
        "token": token,
        "email": req.email,
        "role": user["role"],
        "expires_in": 86400
    }

@app.get("/auth/me")
def me(user=Depends(get_current_user)):
    return user

@app.get("/auth/admins")
def get_admins(user=Depends(get_current_user)):
    """승인자 드롭다운용 admin 목록. 로그인한 사용자라면 조회 가능."""
    users = get_users()
    admins = [email for email, u in users.items() if u.get("role") == "admin"]
    return {"admins": sorted(admins)}

# ══════════════════════════════════════════════════
#  PPV
# ══════════════════════════════════════════════════
@app.get("/ppv")
def get_ppv(user=Depends(get_current_user)):
    return fb_read("ppv")

@app.get("/ppv/summary")
def get_ppv_summary(user=Depends(get_current_user)):
    return fb_read("ppv/summary")

class PPVSummary(BaseModel):
    totalRisked: float
    byStage: dict
    projectCount: int

@app.post("/ppv/summary")
def save_ppv_summary(data: PPVSummary, user=Depends(get_current_user)):
    payload = data.dict()
    payload["updatedAt"] = datetime.datetime.now().isoformat()
    payload["updatedBy"] = user["email"]
    fb_write("ppv/summary", payload)
    return {"ok": True, "data": payload}

@app.post("/ppv/snapshot")
def save_snapshot(data: dict, user=Depends(get_current_user)):
    data["ts"] = datetime.datetime.now().isoformat()
    data["by"] = user["email"]
    requests.post(
        f"{FB_URL}/ppv/snapshots.json",
        json=data,
        params=fb_auth_param(),
        timeout=5
    )
    return {"ok": True}

@app.post("/ppv/event")
def save_event(data: dict, user=Depends(get_current_user)):
    data["ts"] = datetime.datetime.now().isoformat()
    data["by"] = user["email"]
    requests.post(
        f"{FB_URL}/ppv/events.json",
        json=data,
        params=fb_auth_param(),
        timeout=5
    )
    return {"ok": True}

@app.post("/ppv/override/{project_name}")
def save_override(project_name: str, data: dict, user=Depends(get_current_user)):
    safe_name = project_name.replace("/", "_").replace(".", "_")
    data["updatedAt"] = datetime.datetime.now().isoformat()
    data["updatedBy"] = user["email"]
    fb_write(f"ppv/overrides/{safe_name}", data)
    return {"ok": True}

# ══════════════════════════════════════════════════
#  재무 (P&L / B/S / C/F)
# ══════════════════════════════════════════════════
@app.get("/financial")
def get_financial(user=Depends(get_current_user)):
    return fb_read("financial")

@app.get("/financial/{stmt}")
def get_stmt(stmt: str, user=Depends(get_current_user)):
    if stmt not in ["pl", "bs", "cf"]:
        raise HTTPException(status_code=400, detail="stmt는 pl, bs, cf 중 하나여야 합니다.")
    return fb_read(f"financial/{stmt}")

@app.get("/financial/{stmt}/{year}")
def get_stmt_year(stmt: str, year: int, user=Depends(get_current_user)):
    if stmt not in ["pl", "bs", "cf"]:
        raise HTTPException(status_code=400, detail="stmt는 pl, bs, cf 중 하나여야 합니다.")
    return fb_read(f"financial/{stmt}/{year}")

class FinancialData(BaseModel):
    year: int
    month: int
    data: dict

@app.post("/financial/{stmt}")
def save_financial(stmt: str, req: FinancialData, user=Depends(get_current_user)):
    if stmt not in ["pl", "bs", "cf"]:
        raise HTTPException(status_code=400, detail="stmt는 pl, bs, cf 중 하나여야 합니다.")
    payload = req.data.copy()
    payload["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    payload["updatedBy"] = user["email"]
    fb_write(f"financial/{stmt}/{req.year}/{req.month}", payload)
    return {"ok": True, "path": f"financial/{stmt}/{req.year}/{req.month}", "data": payload}

# ══════════════════════════════════════════════════
#  매각 현황
# ══════════════════════════════════════════════════
@app.get("/divest")
def get_divest(user=Depends(get_current_user)):
    return fb_read("divest")

@app.post("/divest/{project_name}")
def update_divest(project_name: str, data: dict, user=Depends(get_current_user)):
    safe_name = project_name.replace("/", "_").replace(".", "_")
    data["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    data["updatedBy"] = user["email"]
    fb_write(f"divest/{safe_name}", data)
    return {"ok": True}

# ══════════════════════════════════════════════════
#  Atlas Milestone
# ══════════════════════════════════════════════════
@app.get("/atlas")
def get_atlas(user=Depends(get_current_user)):
    return fb_read("atlas")

@app.post("/atlas/{milestone_id}")
def update_atlas(milestone_id: str, data: dict, user=Depends(get_current_user)):
    data["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    data["updatedBy"] = user["email"]
    fb_write(f"atlas/{milestone_id}", data)
    return {"ok": True}

# ══════════════════════════════════════════════════
#  전체 데이터 (대시보드 초기 로딩용)
# ══════════════════════════════════════════════════
@app.get("/dashboard")
def get_dashboard(user=Depends(get_current_user)):
    return {
        "ppv_summary": fb_read("ppv/summary"),
        "financial": fb_read("financial"),
        "divest": fb_read("divest"),
        "atlas": fb_read("atlas"),
    }

# ══════════════════════════════════════════════════
#  외부 시장 벤치마크 (FRED + LevelTen)
# ══════════════════════════════════════════════════

# FRED 시리즈 매핑: HEUH 투자 의사결정에 의미있는 지표만
FRED_SERIES = {
    # 금리 (할인율 기준점)
    "us_10y":       {"id": "DGS10",        "label": "US 10Y Treasury",    "unit": "%",       "group": "rates"},
    "us_2y":        {"id": "DGS2",         "label": "US 2Y Treasury",     "unit": "%",       "group": "rates"},
    "fed_funds":    {"id": "DFF",          "label": "Fed Funds Rate",     "unit": "%",       "group": "rates"},
    "bbb_spread":   {"id": "BAMLC0A4CBBB", "label": "BBB Corp Spread",    "unit": "%",       "group": "rates"},
    # 에너지/인플레
    "henry_hub":    {"id": "DHHNGSP",      "label": "Henry Hub NatGas",   "unit": "$/MMBtu", "group": "energy"},
    "cpi":          {"id": "CPIAUCSL",     "label": "US CPI (Index)",     "unit": "Index",   "group": "macro"},
    # 환율
    "krw_usd":      {"id": "DEXKOUS",      "label": "KRW/USD",            "unit": "KRW",     "group": "fx"},
}

# TAN ETF는 FRED에 없음 — Stooq로 별도 조회
STOOQ_SYMBOLS = {
    "tan":          {"symbol": "tan.us",   "label": "TAN (Solar ETF)",    "unit": "$",       "group": "equity"},
    "icln":         {"symbol": "icln.us",  "label": "ICLN (Clean Energy)", "unit": "$",      "group": "equity"},
}

def _fred_fetch(series_id: str, days: int = 180):
    """FRED API에서 시리즈 데이터 조회. 최근 N일치."""
    if not FRED_API_KEY:
        return None
    try:
        end = datetime.date.today()
        start = end - datetime.timedelta(days=days)
        res = requests.get(
            "https://api.stlouisfed.org/fred/series/observations",
            params={
                "series_id": series_id,
                "api_key": FRED_API_KEY,
                "file_type": "json",
                "observation_start": start.isoformat(),
                "observation_end": end.isoformat(),
                "sort_order": "asc",
            },
            timeout=10,
        )
        if res.status_code != 200:
            return None
        obs = res.json().get("observations", [])
        # "." = 결측, 제외
        points = [
            {"date": o["date"], "value": float(o["value"])}
            for o in obs if o.get("value") not in (".", "", None)
        ]
        return points
    except Exception:
        return None

def _stooq_fetch(symbol: str, days: int = 180):
    """Stooq에서 일별 종가 조회 (CSV, 키 불필요)."""
    try:
        url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"
        res = requests.get(url, timeout=10)
        if res.status_code != 200 or "Date,Open" not in res.text:
            return None
        lines = res.text.strip().split("\n")[1:]
        end = datetime.date.today()
        start = end - datetime.timedelta(days=days)
        points = []
        for ln in lines:
            parts = ln.split(",")
            if len(parts) < 5:
                continue
            try:
                d = datetime.date.fromisoformat(parts[0])
                if d < start:
                    continue
                points.append({"date": parts[0], "value": float(parts[4])})
            except Exception:
                continue
        return points
    except Exception:
        return None

def _summarize_series(points):
    """시계열 → 최신값/변동/스파크라인 요약 (1년치)."""
    if not points:
        return None
    latest = points[-1]
    prev = points[-2] if len(points) >= 2 else latest
    # 약 일주일 전 (영업일 5개 전)
    week_ago = points[-6] if len(points) >= 6 else points[0]
    # 약 한달 전 (영업일 21개 전)
    month_ago = points[-22] if len(points) >= 22 else points[0]
    # 약 1년 전 (영업일 252개 전)
    year_ago = points[-253] if len(points) >= 253 else points[0]
    # 1년치 주간 샘플링 (52포인트 내외) — 주 1회 데이터만 추출
    # 영업일 기준 5일마다 1개 선택
    sampled = points[::5] if len(points) > 60 else points
    return {
        "latest": latest["value"],
        "latest_date": latest["date"],
        "d_1d": latest["value"] - prev["value"],
        "d_1w": latest["value"] - week_ago["value"],
        "d_1m": latest["value"] - month_ago["value"],
        "d_1y": latest["value"] - year_ago["value"],
        "spark": [p["value"] for p in sampled],
        "spark_dates": [p["date"] for p in sampled],  # 실제 날짜 병행 전달
        "n_points": len(points),
    }

@app.get("/benchmark/market")
def get_market_benchmark(force: int = 0, user=Depends(get_current_user)):
    """FRED + Stooq 시장 벤치마크. 6시간 캐시."""
    today = datetime.date.today().isoformat()
    cache_key = f"benchmark_cache/market/{today}"

    # 캐시 확인 (force=1이면 무시)
    if not force:
        cached = fb_read(cache_key)
        if cached and cached.get("fetched_at"):
            try:
                fetched = datetime.datetime.fromisoformat(cached["fetched_at"])
                age_hrs = (datetime.datetime.utcnow() - fetched).total_seconds() / 3600
                if age_hrs < 6:
                    return cached
            except Exception:
                pass

    if not FRED_API_KEY:
        raise HTTPException(500, "FRED_API_KEY 환경변수 미설정")

    result = {
        "fetched_at": datetime.datetime.utcnow().isoformat()[:19],
        "source": "FRED + Stooq",
        "series": {},
    }

    # FRED 시리즈 (1년치)
    for key, meta in FRED_SERIES.items():
        pts = _fred_fetch(meta["id"], days=400 if key == "cpi" else 365)
        summary = _summarize_series(pts) if pts else None
        result["series"][key] = {
            **meta,
            "data": summary,
            "ok": summary is not None,
        }

    # Stooq 시리즈 (1년치)
    for key, meta in STOOQ_SYMBOLS.items():
        pts = _stooq_fetch(meta["symbol"], days=365)
        summary = _summarize_series(pts) if pts else None
        result["series"][key] = {
            **meta,
            "data": summary,
            "ok": summary is not None,
        }

    # CPI는 YoY % 변화율로 계산 (Index 자체는 의미가 없음)
    cpi = result["series"].get("cpi", {}).get("data")
    if cpi and cpi.get("spark") and len(cpi["spark"]) >= 13:
        try:
            yoy = (cpi["spark"][-1] / cpi["spark"][-13] - 1) * 100
            result["series"]["cpi"]["yoy_pct"] = round(yoy, 2)
        except Exception:
            pass

    # 캐시 저장
    try:
        fb_put(cache_key, result)
    except Exception:
        pass  # 캐시 실패해도 응답은 반환
    return result


# ── LevelTen PPA Index 업로드/조회 ─────────────────
@app.post("/benchmark/levelten/upload")
async def upload_levelten(
    file: UploadFile = File(...),
    quarter: str = Form(...),  # e.g. "2026-Q1"
    user=Depends(get_current_user),
):
    """LevelTen PPA Index 리포트 업로드 → Claude API로 파싱 → Firebase 저장."""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    raw = await file.read()
    filename = file.filename or "levelten.pdf"
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    # 1) 파일 타입별 텍스트 추출
    source_text = ""
    parse_mode = ""
    parsed = None

    if ext in ("csv", "tsv"):
        try:
            source_text = raw.decode("utf-8", errors="ignore")
        except Exception:
            source_text = raw.decode("latin-1", errors="ignore")
        parse_mode = "csv"

    elif ext in ("xlsx", "xls", "xlsb"):
        parse_mode = "excel"
        tmp_path = None
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}")
            tmp_path = tmp.name
            tmp.write(raw)
            tmp.close()

            if ext == "xlsb":
                lines = []
                with open_workbook(tmp_path) as wb:
                    for sheet_name in wb.sheets[:5]:
                        lines.append(f"\n=== Sheet: {sheet_name} ===")
                        with wb.get_sheet(sheet_name) as sh:
                            for row in sh.rows():
                                vals = [str(c.v) if c.v is not None else "" for c in row]
                                lines.append("\t".join(vals))
                source_text = "\n".join(lines)[:40000]
            else:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    raise HTTPException(400, "openpyxl이 설치되지 않았습니다. requirements.txt에 추가하세요.")
                wb = load_workbook(tmp_path, data_only=True, read_only=True)
                lines = []
                for sheet_name in wb.sheetnames[:5]:
                    lines.append(f"\n=== Sheet: {sheet_name} ===")
                    for row in wb[sheet_name].iter_rows(values_only=True):
                        vals = [str(v) if v is not None else "" for v in row]
                        lines.append("\t".join(vals))
                wb.close()
                source_text = "\n".join(lines)[:40000]
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Excel 파싱 실패: {str(e)}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.unlink(tmp_path)
                except Exception: pass

    elif ext == "pdf":
        parse_mode = "pdf"
        import base64
        pdf_b64 = base64.standard_b64encode(raw).decode("utf-8")

        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        prompt = (
            "You are parsing a LevelTen Energy PPA Price Index report for a Solar+BESS developer. "
            "Extract structured data into strict JSON (no markdown, no prose).\n\n"
            "IMPORTANT PRINCIPLES:\n"
            "1. We focus on Solar PPA and BESS Storage pricing. Skip Wind data entirely.\n"
            "2. ONLY extract values that are explicitly present in the report — tables, charts, or text.\n"
            "3. DO NOT estimate, guess, or use general market knowledge to fill missing values.\n"
            "4. If a value is not in the report, use null. It is BETTER to return null than to invent data.\n"
            "5. For chart-read values (Storage Price Spreads typically shown as charts only), estimate to nearest $0.5 and mark source as 'chart_read'.\n\n"

            "Required schema:\n"
            "{\n"
            '  "quarter": "YYYY-QN",\n'
            '  "report_date": "YYYY-MM-DD or null",\n'
            '  "solar_iso": [\n'
            '    {"region":"ERCOT|PJM|MISO|CAISO|SPP|NYISO|ISO-NE|AESO", "p25":<number $/MWh>,\n'
            '     "qoq_pct":<number or null>, "yoy_pct":<number or null>}\n'
            "  ],\n"
            '  "solar_continental": {\n'
            '    "p25":<number>, "p50":<number>, "p75":<number>,\n'
            '    "p10":<number or null>, "p90":<number or null>,\n'
            '    "qoq_pct":<number or null>, "yoy_pct":<number or null>\n'
            "  },\n"
            '  "solar_hub": [\n'
            '    {"region":"ERCOT", "hub":"HB_NORTH|HB_WEST|HB_SOUTH|HB_HOUSTON|SP15|Alberta|WESTERN HUB|DOM|N ILLINOIS HUB|AEP-DAYTON HUB|SPPNORTH_HUB|SPPSOUTH_HUB|MINN.HUB|ILLINOIS.HUB|INDIANA.HUB|LOUISIANA.HUB|ARKANSAS.HUB|...", "p25":<number>}\n'
            "  ],\n"
            '  "storage_iso": [\n'
            '    {"region":"AESO|CAISO|ERCOT|MISO|PJM|SPP|...", \n'
            '     "min":<number or null>, "p25":<number or null>, "median":<number or null>, "p75":<number or null>, "max":<number or null>,\n'
            '     "source":"levelten_index|chart_read"}\n'
            "  ],\n"
            '  "storage_duration_mix": [\n'
            '    {"region":"ERCOT", "2h":<pct or null>, "3h":<pct or null>, "4h":<pct or null>, "6h":<pct or null>, "8h":<pct or null>, "10h":<pct or null>}\n'
            "  ],\n"
            '  "solar_psv": [\n'
            '    {"region":"ERCOT", "psv_median":<number $/MWh>, "psv_min":<number>, "psv_max":<number>}\n'
            "  ],\n"
            '  "pipeline_breakdown": [\n'
            '    {"cod_year":"2025|2026|2027|2028|2029|2030+", "solar_mw":<number>, "standalone_storage_mw":<number>, "hybrid_mw":<number>}\n'
            "  ],\n"
            '  "storage_available": <true if BESS data found in report, false otherwise>,\n'
            '  "storage_note": "description of BESS data source",\n'
            '  "key_insights": ["1-line insight 1", "1-line insight 2", ...],\n'
            '  "notes": "2-3 sentence summary of quarter trends (Solar + Storage focus)"\n'
            "}\n\n"

            "CRITICAL — Storage Extraction:\n"
            "- LevelTen's 'Storage Price Spreads by ISO' is a BOX PLOT chart showing MIN, P25, MEDIAN, P75, MAX for each ISO.\n"
            "- Read ALL 5 statistics from the box plot. Round to nearest $0.5. Mark source='levelten_index'.\n"
            "- These are LEVELIZED TOLLING AGREEMENT prices in $/kW-month (confirmed by methodology).\n"
            "- Typical ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP (6 ISOs). ISO-NE and NYISO NOT covered by LevelTen Storage Index.\n"
            "- Also extract 'Storage Duration Distribution by ISO' chart → percent for each duration (2h, 3h, 4h, 6h, 8h, 10h).\n\n"

            "CRITICAL — Hub-level Solar P25 Extraction:\n"
            "- Every ISO has a 'PPA Prices by Hub' section showing maps with Solar P25 values labeled on each hub.\n"
            "- Extract every hub + price combination. Examples:\n"
            "  - ERCOT: HB_NORTH, HB_WEST, HB_SOUTH, HB_HOUSTON (4 hubs)\n"
            "  - CAISO: SP15 (1 hub)\n"
            "  - MISO: MINN.HUB, ILLINOIS.HUB, INDIANA.HUB, LOUISIANA.HUB, ARKANSAS.HUB (5 hubs)\n"
            "  - PJM: WESTERN HUB, DOM, AEP-DAYTON HUB, N ILLINOIS HUB (4 hubs)\n"
            "  - SPP: SPPNORTH_HUB, SPPSOUTH_HUB (2 hubs)\n"
            "  - AESO: Alberta (1 hub)\n\n"

            "CRITICAL — Solar PSV (Projected Settlement Value):\n"
            "- Report has 'Projected Settlement Values by Market: Solar' box plot chart.\n"
            "- Read median, min, max for each ISO shown. Values are in $/MWh (can be NEGATIVE).\n"
            "- Typical ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP.\n\n"

            "CRITICAL — Pipeline Breakdown:\n"
            "- Report has 'Technology Breakdown of Pipelines by COD Year' bar chart (in 'Going Hybrid' section).\n"
            "- Extract MW values for each year × technology. Include Solar, Standalone Storage, Hybrid.\n"
            "- DO NOT include Wind.\n\n"

            "General Rules:\n"
            "- Solar prices: USD/MWh. Storage prices: USD/kW-month.\n"
            "- If data not available for a field, use null. NEVER invent numbers.\n"
            "- DO NOT include any Wind data anywhere in the output.\n"
            "- Return ONLY the JSON object. No explanation. No code fences."
        )
        body = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 4000,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    }},
                    {"type": "text", "text": prompt},
                ],
            }],
        }
        try:
            res = requests.post("https://api.anthropic.com/v1/messages",
                                headers=headers, json=body, timeout=150)
            if res.status_code != 200:
                raise HTTPException(502, f"Claude API 오류: {res.text[:300]}")
            ai_text = res.json()["content"][0]["text"].strip()
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```", 2)[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
                ai_text = ai_text.rsplit("```", 1)[0]
            parsed = json.loads(ai_text.strip())
        except HTTPException:
            raise
        except requests.Timeout:
            raise HTTPException(504, "Claude API 응답 타임아웃 — 리포트가 너무 크거나 서버 혼잡 (재시도 권장)")
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI 응답 JSON 파싱 실패: {str(e)}")
    else:
        raise HTTPException(400, f"지원하지 않는 파일 형식: .{ext} (PDF, CSV, XLSX, XLSB 지원)")

    # CSV/Excel인 경우 Claude에게 텍스트 파싱 요청
    if parse_mode in ("csv", "excel"):
        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        prompt = (
            "You are parsing LevelTen Energy PPA Price Index tabular data for a Solar+BESS developer. "
            f"The data below is from a {parse_mode.upper()} export.\n\n"
            f"DATA:\n{source_text[:30000]}\n\n"
            "IMPORTANT PRINCIPLES:\n"
            "1. We focus on Solar PPA and BESS Storage pricing. Skip Wind data.\n"
            "2. ONLY extract values that are EXPLICITLY present in the data. NEVER estimate or invent.\n"
            "3. If a value is missing, use null. Do NOT fill with guesses.\n\n"
            "Extract into strict JSON (no markdown, no prose):\n"
            "{\n"
            '  "quarter": "YYYY-QN",\n'
            '  "report_date": "YYYY-MM-DD or null",\n'
            '  "solar_iso": [{"region":"ERCOT|PJM|MISO|CAISO|SPP|ISO-NE|AESO", "p25":<$/MWh>, "qoq_pct":<null>, "yoy_pct":<null>}],\n'
            '  "solar_continental": {"p25":<number>, "p50":<number>, "p75":<number>, "p10":<number|null>, "p90":<number|null>, "qoq_pct":<null>, "yoy_pct":<null>},\n'
            '  "solar_hub": [{"region":"ERCOT", "hub":"North", "p25":<number>}],\n'
            '  "storage_iso": [{"region":"ERCOT|...", "p25":<$/kW-month|null>, "p50":<number|null>, "p75":<number|null>, "source":"table"}],\n'
            '  "storage_available": <true|false>,\n'
            '  "storage_note": "description or \\"Not included in data\\"",\n'
            '  "key_insights": ["actionable insight 1", "insight 2"],\n'
            '  "notes": "2-3 sentence summary (Solar+Storage focus, no Wind)"\n'
            "}\n"
            "If NO storage data in file: storage_iso=[], storage_available=false.\n"
            "Return ONLY the JSON object. No code fences."
        )
        body = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 4000,
            "messages": [{"role": "user", "content": prompt}],
        }
        try:
            res = requests.post("https://api.anthropic.com/v1/messages",
                                headers=headers, json=body, timeout=150)
            if res.status_code != 200:
                raise HTTPException(502, f"Claude API 오류: {res.text[:300]}")
            ai_text = res.json()["content"][0]["text"].strip()
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```", 2)[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
                ai_text = ai_text.rsplit("```", 1)[0]
            parsed = json.loads(ai_text.strip())
        except HTTPException:
            raise
        except requests.Timeout:
            raise HTTPException(504, "Claude API 응답 타임아웃 — 리포트가 너무 크거나 서버 혼잡 (재시도 권장)")
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"AI 응답 JSON 파싱 실패: {str(e)}")

    if not parsed:
        raise HTTPException(500, "파싱 결과가 비어있습니다.")

    # 쿼터 형식 검증 (YYYY-QN)
    import re
    if not re.match(r'^\d{4}-Q[1-4]$', quarter.upper()):
        raise HTTPException(400, "쿼터 형식은 YYYY-Q1 ~ YYYY-Q4 여야 합니다.")
    quarter = quarter.upper()

    # 쿼터 덮어쓰기 (사용자가 명시한 값이 우선)
    parsed["quarter"] = quarter
    parsed["uploaded_at"] = datetime.datetime.utcnow().isoformat()[:19]
    parsed["uploaded_by"] = user["email"]
    parsed["filename"] = filename
    parsed["parse_mode"] = parse_mode

    # Backward compat: 새 파서 스키마를 legacy entries 배열로도 변환
    if "entries" not in parsed:
        legacy = []
        for s in parsed.get("solar_iso", []) or []:
            legacy.append({"tech":"solar", "region": s.get("region",""), "term_yr":10,
                           "p25": s.get("p25"), "p50": None, "p75": None})
        for s in parsed.get("storage_iso", []) or []:
            # 새 스키마: min/p25/median/p75/max → legacy: p25/p50/p75
            legacy.append({"tech":"storage", "region": s.get("region",""), "term_yr":10,
                           "p25": s.get("p25"),
                           "p50": s.get("median") or s.get("p50"),
                           "p75": s.get("p75")})
        parsed["entries"] = legacy

    # Firebase 저장: benchmark/levelten/{quarter}
    fb_put(f"benchmark/levelten/{quarter}", parsed)

    return {
        "ok": True,
        "quarter": quarter,
        "solar_iso_count": len(parsed.get("solar_iso", []) or []),
        "storage_iso_count": len(parsed.get("storage_iso", []) or []),
        "entries_count": len(parsed.get("entries", []) or []),
        "data": parsed,
    }


@app.get("/benchmark/levelten")
def get_levelten_all(user=Depends(get_current_user)):
    """모든 분기별 LevelTen 데이터."""
    return fb_read("benchmark/levelten") or {}


@app.get("/benchmark/levelten/latest")
def get_levelten_latest(user=Depends(get_current_user)):
    """가장 최신 분기의 LevelTen 데이터."""
    all_data = fb_read("benchmark/levelten") or {}
    if not all_data:
        return {}
    # 쿼터 문자열 정렬 (YYYY-QN 포맷이라 사전순 정렬로 충분)
    latest_key = sorted(all_data.keys())[-1]
    return {"quarter": latest_key, **all_data[latest_key]}


@app.delete("/benchmark/levelten/{quarter}")
def delete_levelten(quarter: str, user=Depends(require_admin)):
    """특정 분기 LevelTen 데이터 삭제."""
    import re
    if not re.match(r'^\d{4}-Q[1-4]$', quarter.upper()):
        raise HTTPException(400, "쿼터 형식은 YYYY-Q1 ~ YYYY-Q4 여야 합니다.")
    quarter = quarter.upper()
    try:
        requests.delete(f"{FB_URL}/benchmark/levelten/{quarter}.json",
                        params=fb_auth_param(), timeout=5)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, f"삭제 오류: {str(e)}")


# ── 피어 IRR 벤치마크 (내부 수동 입력값) ────────────
@app.get("/benchmark/peer-irr")
def get_peer_irr(user=Depends(get_current_user)):
    """저장된 피어 IRR 벤치마크 조회."""
    return fb_read("benchmark/peer_irr") or {}


@app.post("/benchmark/peer-irr")
def save_peer_irr(payload: dict, user=Depends(get_current_user)):
    """피어 IRR 벤치마크 저장 (Levered Pre-Tax IRR 레인지)."""
    # 필드 검증
    required_numeric = ["solar_min", "solar_max", "hybrid_min", "hybrid_max", "wind_min", "wind_max"]
    data = {}
    for k in required_numeric:
        v = payload.get(k)
        if v is None:
            raise HTTPException(400, f"누락된 필드: {k}")
        try:
            fv = float(v)
            if fv < 0 or fv > 50:
                raise HTTPException(400, f"{k}: 0~50% 범위여야 합니다.")
            data[k] = round(fv, 2)
        except (ValueError, TypeError):
            raise HTTPException(400, f"{k}: 숫자여야 합니다.")
    # min < max 검증
    for tech in ("solar", "hybrid", "wind"):
        if data[f"{tech}_min"] >= data[f"{tech}_max"]:
            raise HTTPException(400, f"{tech}: min < max 이어야 합니다.")
    # 비고
    note = payload.get("note", "")
    if isinstance(note, str):
        data["note"] = note[:200]
    data["updated_at"] = datetime.datetime.utcnow().isoformat()[:19]
    data["updated_by"] = user["email"]
    fb_put("benchmark/peer_irr", data)
    return {"ok": True, "data": data}


# ══════════════════════════════════════════════════
#  BESS Tolling Market Research (AI Web Search)
# ══════════════════════════════════════════════════
@app.post("/benchmark/bess-tolling/research")
def research_bess_tolling(user=Depends(get_current_user)):
    """
    Claude API + web_search 도구로 ISO별 BESS tolling 가격을 실시간 리서치.
    결과: ISO × Duration 별 P25/P75 + 출처 URL + confidence score.
    캐시: benchmark/bess_tolling/latest (수동 새로고침)
    """
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 미설정")

    today_str = datetime.date.today().isoformat()

    prompt = (
        f"You are an energy market research analyst specializing in US battery energy storage "
        f"system (BESS) tolling agreements AND PPA markets. Today: {today_str}.\n\n"
        "ROLE: This research is COMPLEMENTARY to LevelTen's official PPA Price Index.\n"
        "LevelTen publishes official data for 6 ISOs: AESO, CAISO, ERCOT, MISO, PJM, SPP.\n"
        "For these 6 ISOs the dashboard uses LevelTen first — your role is DURATION-level BESS detail only.\n\n"
        "YOUR FOCUS (three objectives):\n"
        "  (A) NON-LEVELTEN ISOs — provide BOTH BESS tolling AND PPA market commentary:\n"
        "      - ISO-NE (New England)\n"
        "      - NYISO (New York)\n"
        "      - WECC_DSW (Desert Southwest: AZ, NM, NV — Arizona/New Mexico/Nevada utilities)\n"
        "      - WECC_RM  (Rocky Mountain: UT, CO, WY, ID — PacifiCorp East/RMP, Xcel Colorado)\n"
        "      - WECC_NW  (Northwest: OR, WA, MT — PacifiCorp West, PGE, Puget Sound Energy)\n"
        "      - SERC (Southeast: TVA, Duke, Southern Company territory — NC, SC, GA, AL, TN, KY)\n"
        "  (B) DURATION BREAKDOWN for LevelTen-covered ISOs (ERCOT, CAISO, PJM, MISO, SPP, AESO):\n"
        "      → duration-level prices (2h / 4h / 6h) — LevelTen only gives ISO-level\n"
        "  (C) For WECC sub-regions: include PPA market commentary since LevelTen has ZERO coverage.\n"
        "      Key utility RFPs to reference: PacifiCorp IRP RFP, URC (Utah Renewable Communities),\n"
        "      APS (Arizona Public Service), NV Energy, Xcel Energy Colorado, Portland General Electric,\n"
        "      Idaho Power, Puget Sound Energy.\n\n"
        "Research methodology — TRIANGULATION:\n"
        "  1. Capacity market clearing prices (PJM, NYISO, ISO-NE) — adjusted for storage\n"
        "  2. Merchant BESS revenue data (ERCOT ~$30-50/kW-yr, CAISO duck curve premium)\n"
        "  3. Utility RFP announcements when prices are disclosed (PacifiCorp, APS, Xcel, etc.)\n"
        "  4. Public company earnings calls (NextEra, Vistra, AES)\n"
        "  5. Duration-adjustment heuristic:\n"
        "     - 2h: 60-75% of 4h price (arbitrage-dominated)\n"
        "     - 4h: reference (capacity-dominated, NERC/ISO standard)\n"
        "     - 6h+: 110-130% of 4h (long-duration premium)\n"
        "  6. Industry rule-of-thumb benchmarks (2025):\n"
        "     - ERCOT 2h: $3-8/kW-mo  | 4h: $5-12/kW-mo\n"
        "     - CAISO 4h: $10-16/kW-mo (duck curve) | 8h: $13-20/kW-mo\n"
        "     - PJM 4h: $8-13/kW-mo (capacity market) | 2h: $5-9/kW-mo\n"
        "     - SPP/MISO 4h: $6-11/kW-mo\n"
        "     - ISO-NE 4h: $12-18/kW-mo (tight capacity, winter peak)\n"
        "     - NYISO 4h: $11-17/kW-mo (DEC mandate, expensive zones J/K)\n"
        "     - WECC_DSW 4h: $7-12/kW-mo (APS/NV Energy RFPs, solar-shifting demand)\n"
        "     - WECC_RM 4h: $6-11/kW-mo (PacifiCorp/Xcel CO — emerging market, thin liquidity)\n"
        "     - WECC_NW 4h: $6-10/kW-mo (hydro-dominant, moderate storage need)\n"
        "     - SERC 4h: $7-12/kW-mo (vertically integrated utilities, bilateral)\n"
        "Use these as STARTING POINTS, then VERIFY/ADJUST via web_search.\n\n"
        "Use web_search to find CURRENT data from:\n"
        "- Wood Mackenzie, BloombergNEF, S&P Global, LCG Consulting\n"
        "- ISO capacity auction results: PJM BRA, ISO-NE FCA, NYISO ICAP\n"
        "- State PUC filings for RFP results (Utah PSC, Colorado PUC, Oregon PUC, Arizona ACC)\n"
        "- Utility IRP documents (PacifiCorp IRP, APS IRP, Xcel Colorado ERP)\n"
        "- Press releases: NextEra, Invenergy, AES, EDP, Engie, Brookfield\n"
        "- News: Utility Dive, Energy Storage News, Reuters, Canary Media\n\n"
        "TARGET REGIONS (10 total):\n"
        "- PRIMARY (no LevelTen coverage, full research required):\n"
        "    ISO-NE, NYISO, WECC_DSW, WECC_RM, WECC_NW, SERC\n"
        "- SECONDARY (LevelTen-covered, provide duration breakdown only):\n"
        "    ERCOT, CAISO, PJM, MISO, SPP, AESO\n\n"
        "TARGET DURATIONS for each region: 2h, 4h, 6h\n\n"
        "Output: ALL text fields (market_note, methodology_note, caveats) MUST BE IN KOREAN.\n"
        "Use formal nominal/concise style ('~확인됨', '~추정됨', '~범위').\n"
        "For WECC_* regions, market_note MUST include PPA market commentary (utility RFP landscape, "
        "recent clearing prices, Neptune-like Utah projects context).\n"
        "Numbers stay numeric ($X/kW-mo). Region names stay English.\n\n"
        "Return ONLY this JSON structure (no markdown, no code fences):\n"
        "{\n"
        '  "research_date": "YYYY-MM-DD",\n'
        '  "iso_data": [\n'
        '    {\n'
        '      "region": "ERCOT|CAISO|PJM|MISO|SPP|AESO|ISO-NE|NYISO|WECC_DSW|WECC_RM|WECC_NW|SERC",\n'
        '      "levelten_covered": true,  // 6 LevelTen ISOs=true, 나머지 4 (ISO-NE/NYISO/WECC_*/SERC)=false\n'
        '      "durations": [\n'
        '        {"hours": 2, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"},\n'
        '        {"hours": 4, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"},\n'
        '        {"hours": 6, "p25": <number>, "p75": <number>, "confidence": "high|medium|low"}\n'
        '      ],\n'
        '      "market_note": "(한국어) 시장 특성 1-2문장. WECC_*는 PPA 시장 commentary 포함 (주요 utility RFP, recent clearing prices, 인접 주 벤치마크)",\n'
        '      "sources": [\n'
        '        {"url": "https://...", "title": "source title", "date": "YYYY-MM", "key_data": "핵심 수치/인용 (한국어 번역 OK)"}\n'
        '      ]\n'
        '    }\n'
        '  ],\n'
        '  "methodology_note": "(한국어) 추정 방법 요약. LevelTen 공식 index와의 관계 명시: LevelTen은 6개 ISO만 커버, 본 리서치는 (1) 미커버 4개 지역(ISO-NE/NYISO/WECC_DSW/WECC_RM/WECC_NW/SERC) 보완, (2) 전체 duration별(2h/4h/6h) 세분화 목표. capacity market + merchant 수익 + utility RFP 삼각 검증",\n'
        '  "confidence_overall": "high|medium|low",\n'
        '  "caveats": "(한국어) 1-2문장. 예: 본 수치는 AI 리서치 기반 추정치. LevelTen 6개 ISO는 공식 데이터 우선. WECC sub-region 및 SERC는 공식 index 부재 — RFP/IRP 참고치"\n'
        "}\n\n"
        "Rules:\n"
        "- All prices in USD/kW-month, levelized over contract term.\n"
        "- ALWAYS include all 12 regions (10 + WECC split into 3 sub-regions):\n"
        "  ERCOT, CAISO, PJM, MISO, SPP, AESO, ISO-NE, NYISO, WECC_DSW, WECC_RM, WECC_NW, SERC.\n"
        "- Each region must have 3 durations (2h, 4h, 6h) — use benchmark if no evidence, mark confidence='low'.\n"
        "- Confidence guide: 'high' if 3+ sources corroborate; 'medium' if 1-2 sources; 'low' if benchmark/inference only.\n"
        "- For WECC_* regions, market_note MUST include PPA context (not just BESS) — target utilities and recent RFP clearing prices.\n"
        "- Dates must be 2024-2026 (recent only).\n"
        "- All text fields in Korean formal nominal style.\n"
        "- Return valid JSON only."
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-5",
                "max_tokens": 8000,
                "tools": [{
                    "type": "web_search_20250305",
                    "name": "web_search",
                    "max_uses": 10,
                }],
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=180,  # 웹서치 여러 번 → 최대 3분
        )
        if resp.status_code != 200:
            raise HTTPException(502, f"Claude API 오류: {resp.text[:400]}")

        data = resp.json()
        # content blocks 중 text 타입만 합쳐서 JSON 파싱
        text_parts = []
        for block in data.get("content", []):
            if block.get("type") == "text":
                text_parts.append(block.get("text", ""))
        full_text = "".join(text_parts).strip()

        # JSON 추출 (code fence 제거 + { } 범위)
        import re as _re
        clean = _re.sub(r"```(?:json)?\s*", "", full_text).strip().strip("`")
        start = clean.find("{")
        end = clean.rfind("}") + 1
        if start < 0 or end <= start:
            raise HTTPException(500, f"AI 응답에서 JSON을 찾을 수 없음: {full_text[:300]}")
        clean = clean[start:end]

        try:
            parsed = json.loads(clean)
        except json.JSONDecodeError as e:
            raise HTTPException(500, f"JSON 파싱 실패: {str(e)}. 응답: {clean[:400]}")

        # 메타데이터 추가
        parsed["generated_at"] = datetime.datetime.utcnow().isoformat()[:19]
        parsed["generated_by"] = user["email"]
        parsed["source"] = "ai_research"
        # 토큰 사용량 (cost 추적용)
        usage = data.get("usage", {})
        parsed["tokens"] = {
            "input": usage.get("input_tokens", 0),
            "output": usage.get("output_tokens", 0),
        }

        # Firebase 저장 (latest로)
        fb_put("benchmark/bess_tolling/latest", parsed)
        # 히스토리도 (월별 캐시)
        month_key = datetime.date.today().strftime("%Y-%m")
        fb_put(f"benchmark/bess_tolling/history/{month_key}", parsed)

        return {"ok": True, "data": parsed}

    except HTTPException:
        raise
    except requests.Timeout:
        raise HTTPException(504, "AI 리서치 타임아웃 (3분 초과)")
    except Exception as e:
        raise HTTPException(500, f"BESS 리서치 실패: {str(e)}")


@app.get("/benchmark/bess-tolling")
def get_bess_tolling(user=Depends(get_current_user)):
    """저장된 BESS tolling 리서치 결과 조회 (latest)."""
    return fb_read("benchmark/bess_tolling/latest") or {}


@app.get("/benchmark/bess-tolling/history")
def get_bess_tolling_history(user=Depends(get_current_user)):
    """월별 히스토리 (stale 확인용)."""
    return fb_read("benchmark/bess_tolling/history") or {}


# ══════════════════════════════════════════════════
#  Valuation Calculate (Stage 1 Engine)
# ══════════════════════════════════════════════════
import numpy as np
import numpy_financial as npf

def _calc_engine(inputs: dict) -> dict:
    pv_mwac   = inputs.get('pv_mwac', 199)
    pv_mwdc   = pv_mwac * inputs.get('dc_ac_ratio', 1.34)
    bess_mw   = inputs.get('bess_mw', 199)
    bess_mwh  = inputs.get('bess_mwh', 796)
    life      = int(inputs.get('life', 35))

    # CAPEX
    module_cwp   = inputs.get('module_cwp', 31.5)
    bos_cwp      = inputs.get('bos_cwp', 42.9)
    ess_per_kwh  = inputs.get('ess_per_kwh', 234.5)
    epc_cont_pct = inputs.get('epc_cont_pct', 8.0)
    owner_pct    = inputs.get('owner_pct', 3.0)
    softcost_pct = inputs.get('softcost_pct', 5.0)
    intercon_m   = inputs.get('intercon_m', 120.0)
    dev_cost_m   = inputs.get('dev_cost_m', 20.0)
    capex_etc    = inputs.get('capex_etc', 0)

    pv_module = pv_mwdc*1000*module_cwp/100
    pv_bos    = pv_mwdc*1000*bos_cwp/100
    ess_equip = bess_mwh*ess_per_kwh
    epc_base  = pv_module + pv_bos + ess_equip
    epc_total = epc_base * (1 + epc_cont_pct/100)
    pre_capex = (epc_total*(1+owner_pct/100+softcost_pct/100)
                 + intercon_m*1000 + dev_cost_m*1000 + capex_etc*1000)
    int_rate   = inputs.get('int_rate', 5.5) / 100
    debt_ratio = inputs.get('debt_ratio', 47.6) / 100
    base_capex = pre_capex * (1 + debt_ratio*int_rate*0.75 + 0.012)
    total_capex = float(inputs['capex_total_override'])*1000 if inputs.get('capex_total_override') else base_capex

    dev_margin  = pv_mwac*1000*inputs.get('dev_margin_kwac', 200)/1000
    epc_margin  = epc_base * inputs.get('epc_margin_pct', 7.95)/100
    total_margin = dev_margin + epc_margin

    loan_term  = int(inputs.get('loan_term', 18))
    debt       = total_capex * debt_ratio
    ann_ds     = float(npf.pmt(int_rate, loan_term, -debt)) if debt > 0 else 0

    # TE Flip
    _fy_raw = inputs.get('flip_yield', 8.75)
    if _fy_raw > 50: _fy_raw = _fy_raw / 100   # 875 → 8.75 자동 보정
    flip_yield = _fy_raw / 100
    flip_term  = int(inputs.get('flip_term', 7))
    itc_elig   = inputs.get('itc_elig', 97) / 100
    itc_rate   = inputs.get('itc_rate') or inputs.get('credit_val', 30)
    itc_rate   = itc_rate / 100
    te_mult    = inputs.get('te_mult', 1.115)
    yield_adj  = 1 / (1 + (flip_yield - 0.0875) * 8)

    # ── TE Invest 산정 + Sponsor Equity 최소선 확보 ──────────────────
    # 1) 이론치: ITC 기반 TE 투자 규모
    te_theoretical = total_capex * itc_elig * itc_rate * te_mult * yield_adj

    # 2) 구조적 상한: Debt + TE + Sponsor Eq = CAPEX, Sponsor Eq는 최소 10% CAPEX 확보
    #    (Sponsor Eq <5% 되면 IRR 계산 왜곡 — Sponsor 기여도가 미미해서 IRR 발산)
    min_sponsor_eq_pct = inputs.get('min_sponsor_eq_pct', 10.0) / 100  # 기본 10% of CAPEX
    max_te_invest = total_capex - debt - total_capex * min_sponsor_eq_pct

    # 3) 실제 TE 투자: 이론치와 구조적 상한 중 작은 값
    te_invest = max(0, min(te_theoretical, max_te_invest))

    sponsor_eq = total_capex - debt - te_invest
    effective_eq = sponsor_eq * (1 - int_rate * 0.75)

    # MACRS depreciation
    tax_rate    = inputs.get('tax_rate', 21) / 100
    macrs_basis = total_capex * itc_elig * (1 - itc_rate/2)
    depr_sched  = {i+1: macrs_basis*r for i,r in enumerate(MACRS_5YR)}
    depr_share  = inputs.get('depr_share', 0.7721)  # calibrated to Neptune

    # Cash allocation
    pre_flip_cash_te  = inputs.get('pre_flip_cash_te', 25) / 100
    post_flip_cash_te = inputs.get('post_flip_cash_te', 5) / 100

    # Revenue
    cf_pct       = inputs.get('cf_pct', 21.24)
    net_prod_yr1 = inputs.get('net_prod_yr1', None)
    ann_prod_yr1 = float(net_prod_yr1) if net_prod_yr1 else pv_mwac*cf_pct/100*8760
    ppa_price   = inputs.get('ppa_price', 68.82)
    ppa_term    = int(inputs.get('ppa_term', 25))
    ppa_esc     = inputs.get('ppa_esc', 0) / 100
    # bess_toll: CF_Annual Y1 실제값 우선, 없으면 Summary 파싱값
    bess_toll   = inputs.get('bess_toll_y1_effective') or inputs.get('bess_toll', 14.50)
    bess_toll_t = int(inputs.get('bess_toll_term', 20))
    bess_toll_esc = inputs.get('bess_toll_esc', 0) / 100  # Toll escalation (%)
    merch_ppa   = inputs.get('merchant_ppa', 45.0)
    merch_esc   = inputs.get('merchant_esc', 3.0) / 100
    degradation = inputs.get('degradation', 0.0064)
    avail_1     = inputs.get('availability_yr1', 0.977)
    avail_2     = inputs.get('availability_yr2', 0.982)

    # OPEX
    pv_om=inputs.get('pv_om',4.5); pv_om_nc=inputs.get('pv_om_nc',1.0)
    pv_aux=inputs.get('pv_aux',1.56); bess_om=inputs.get('bess_om',8.64)
    bess_om_nc=inputs.get('bess_om_nc',1.0); bess_aux=inputs.get('bess_aux',3.84)
    ins_pv=inputs.get('insurance_pv',10.57); ins_bess=inputs.get('insurance_bess',5.05)
    asset_mgmt=inputs.get('asset_mgmt',210); prop_tax=inputs.get('prop_tax_yr1',3162)
    land_rent=inputs.get('land_rent_yr1',437); opex_etc=inputs.get('opex_etc',0)
    opex_esc=inputs.get('opex_esc',2.0)/100

    # Augmentation
    aug_price=inputs.get('aug_price',150); aug_mwh_pct=inputs.get('aug_mwh_pct',18.8)
    aug_mwh_ea=bess_mwh*aug_mwh_pct/100
    aug_years=[int(y) for y in [inputs.get('aug_y1',4),inputs.get('aug_y2',8),inputs.get('aug_y3',12)] if y and int(y)>0]
    aug_cost_ea=aug_mwh_ea*aug_price

    # Full 35-year CF schedule
    cashflows=[-effective_eq]; unlev_cfs=[-total_capex]; sponsor_cfs=[-effective_eq]; pretax_cfs=[-effective_eq]
    debt_bal=debt; detail=[]; ebitda_yr1=None

    for yr in range(1, life+1):
        avail = avail_1 if yr==1 else avail_2
        prod  = ann_prod_yr1 * avail * ((1-degradation)**(yr-1))

        # CF_Annual parsed schedule 우선 사용 (실제 Neptune 모델값)
        pv_sched   = inputs.get('pv_rev_schedule', [])
        bess_sched = inputs.get('bess_rev_schedule', [])
        merch_sched= inputs.get('merch_rev_schedule', [])

        if pv_sched and yr-1 < len(pv_sched):
            pv_rev = pv_sched[yr-1]
        elif yr <= ppa_term:
            pv_rev = prod*ppa_price*((1+ppa_esc)**(yr-1))/1000
        else:
            pv_rev = prod*merch_ppa*((1+merch_esc)**(yr-1))/1000

        if bess_sched and yr-1 < len(bess_sched):
            bess_rev = bess_sched[yr-1]
        else:
            bess_rev = bess_mw*1000*bess_toll*((1+bess_toll_esc)**(yr-1))*12/1000 if yr<=bess_toll_t else 0

        if merch_sched and yr-1 < len(merch_sched) and merch_sched[yr-1] > 0:
            pv_rev = merch_sched[yr-1]  # merchant 기간은 merch_sched 우선

        total_rev = pv_rev + bess_rev

        esc=(1+opex_esc)**(yr-1); prop_esc=max(0.35,1-0.025*(yr-1))
        opex=(pv_mwdc*1000*pv_om/1000*esc + pv_mwac*1000*pv_om_nc/1000*esc +
              pv_mwac*1000*pv_aux/1000*esc + bess_mw*1000*bess_om/1000*esc +
              bess_mw*1000*bess_om_nc/1000*esc + bess_mw*1000*bess_aux/1000*esc +
              pv_mwac*1000*ins_pv/1000*esc + bess_mw*1000*ins_bess/1000*esc +
              asset_mgmt*esc + prop_tax*prop_esc + land_rent*esc + opex_etc*1000*esc)

        ebitda = total_rev - opex
        if yr==1: ebitda_yr1=ebitda
        aug_c = aug_cost_ea if yr in aug_years else 0

        if yr<=loan_term and debt_bal>0:
            int_p=debt_bal*int_rate; prin=ann_ds-int_p
            ds=ann_ds; debt_bal=max(0,debt_bal-prin)
        else: ds=0

        # MACRS depreciation tax benefit to sponsor
        depr = depr_sched.get(yr, 0)
        s_tax = depr * tax_rate * depr_share  # sponsor keeps depr_share of tax benefit

        op_cf = ebitda - ds - aug_c
        if yr<=flip_term:
            s_cf = op_cf*(1-pre_flip_cash_te) + s_tax
        else:
            s_cf = op_cf*(1-post_flip_cash_te) + s_tax

        s_cf_pretax = op_cf*(1-pre_flip_cash_te) if yr<=flip_term else op_cf*(1-post_flip_cash_te)
        cashflows.append(op_cf); unlev_cfs.append(ebitda-aug_c); sponsor_cfs.append(s_cf); pretax_cfs.append(s_cf_pretax)
        if yr<=10:
            detail.append({'yr':yr,'rev':round(total_rev,0),'opex':round(opex,0),
                'ebitda':round(ebitda,0),'ds':round(ds,0),'aug':round(aug_c,0),
                'depr':round(depr,0),'s_cf':round(s_cf,0)})

    lirr = _irr_robust(pretax_cfs, guess=0.10)   # Sponsor pretax levered (Neptune Row 26 ~10%)
    uirr = _irr_robust(unlev_cfs, guess=0.05)    # Asset-level unlevered (Neptune Row 27 ~8%)
    sirr = _irr_robust(sponsor_cfs, guess=0.10)  # Sponsor after-tax w/ MACRS (Full Life)
    sirr_c = float(npf.irr(sponsor_cfs[:ppa_term+1]))
    ebitda_yield = ebitda_yr1/total_capex*100 if total_capex else 0

    # ── NPV 계산 (Hurdle 기준 할인) ────────────────────────────────
    # Sponsor NPV: Hurdle IRR(예: 10%)로 할인 — 매수자 관점 가치
    # Project NPV: WACC로 할인 — 프로젝트 자체 가치
    hurdle_sponsor = inputs.get('hurdle_sponsor_irr', 10.0) / 100  # Default 10%
    # WACC 계산 (approximation): tax-adjusted weighted cost
    wacc_debt_cost = int_rate * (1 - tax_rate)  # after-tax
    wacc_te_cost = 0.07   # TE 조달 비용 (typical)
    wacc_eq_cost = 0.11   # Sponsor eq 비용 (typical)
    debt_w = debt / total_capex if total_capex else 0
    te_w = te_invest / total_capex if total_capex else 0
    eq_w = sponsor_eq / total_capex if total_capex else 0
    wacc = (debt_w * wacc_debt_cost) + (te_w * wacc_te_cost) + (eq_w * wacc_eq_cost)
    if wacc <= 0 or wacc > 0.5: wacc = 0.072  # fallback

    def _npv(cfs, rate):
        try:
            return float(npf.npv(rate, cfs))
        except Exception:
            return None

    sponsor_npv = _npv(sponsor_cfs, hurdle_sponsor)
    project_npv = _npv(unlev_cfs, wacc)
    # ───────────────────────────────────────────────────────────────

    return {
        'capex_total':   round(total_capex,0),
        'epc_base':      round(epc_base,0),
        'debt':          round(debt,0),
        'equity':        round(sponsor_eq+te_invest,0),
        'te_invest':     round(te_invest,0),
        'sponsor_equity':round(sponsor_eq,0),
        'dev_margin':    round(dev_margin,0),
        'epc_margin':    round(epc_margin,0),
        'total_margin':  round(total_margin,0),
        'levered_irr':   round(lirr,6) if not np.isnan(lirr) else None,
        'unlevered_irr': round(uirr,6) if not np.isnan(uirr) else None,
        'sponsor_irr':   round(sirr,6) if not np.isnan(sirr) else None,
        'sponsor_irr_contract': round(sirr_c,6) if not np.isnan(sirr_c) else None,
        'sponsor_npv':   round(sponsor_npv,0) if sponsor_npv is not None else None,
        'project_npv':   round(project_npv,0) if project_npv is not None else None,
        'wacc':          round(wacc,6),
        'hurdle_sponsor_irr_used': round(hurdle_sponsor,6),
        'ebitda_yield':  round(ebitda_yield,2),
        'aug_cost_ea':   round(aug_cost_ea,0),
        'annual_detail': detail,
        'cashflows':     [round(x,0) for x in cashflows[:36]],
    }


@app.post("/valuation/calculate")
def calculate_valuation(req: ValuationCalcRequest, user=Depends(get_current_user)):
    """Run PF calculation engine with given inputs"""
    try:
        result = _calc_engine(req.inputs)
        return {"ok": True, "project_id": req.project_id, "result": result}
    except Exception as e:
        raise HTTPException(500, f"Calculation error: {str(e)}")


# ── Break-Even Analysis (Newton-Raphson) ─────────
class BreakEvenRequest(BaseModel):
    project_id: str = ""
    inputs: dict
    target_irr_pct: float  # e.g., 11.0 for 11%
    target_var: str = "ppa_price"  # 현재는 PPA만 지원 (확장 가능)

@app.post("/valuation/breakeven")
def break_even(req: BreakEvenRequest, user=Depends(get_current_user)):
    """
    Newton-Raphson 기반 정확한 PPA 역산.
    Phase 1: PPA ±25% 11점 민감도 스캔
    Phase 2: Newton-Raphson (tolerance 0.01% IRR)
    """
    try:
        base_inputs = dict(req.inputs)
        base_ppa = float(base_inputs.get("ppa_price", 68.82))
        target_irr = req.target_irr_pct / 100.0  # 0.11
        tol = 0.0001  # 0.01% IRR tolerance
        h = 0.50  # finite difference step $/MWh
        max_iter = 10

        def calc_irr(ppa_val):
            """Calc engine 호출 → Sponsor IRR (Full Life 우선) 반환"""
            inp = dict(base_inputs)
            inp["ppa_price"] = ppa_val
            res = _calc_engine(inp)
            return res.get("sponsor_irr") or res.get("sponsor_irr_contract") or 0.0

        # ── Phase 1: ±25% 민감도 스캔 ──
        pcts = [-25, -20, -15, -10, -5, 0, 5, 10, 15, 20, 25]
        sensitivity = []
        for pct in pcts:
            ppa_p = base_ppa * (1 + pct / 100.0)
            irr_p = calc_irr(ppa_p)
            sensitivity.append({
                "pct": pct,
                "ppa": round(ppa_p, 2),
                "irr_pct": round(irr_p * 100, 4)
            })

        # ── Phase 2: Newton-Raphson ──
        iterations = []
        ppa = base_ppa  # 초기값
        status = "not_started"
        solution = None

        # 가능 여부 체크: ±25% 범위에 Target이 있는지
        min_irr = sensitivity[0]["irr_pct"]
        max_irr = sensitivity[-1]["irr_pct"]
        target_irr_pct = target_irr * 100

        if target_irr_pct < min_irr or target_irr_pct > max_irr:
            # 범위 밖: 가장 가까운 끝 PPA로 시작 (그래도 시도)
            if target_irr_pct < min_irr:
                ppa = base_ppa * 0.75
                status = "target_below_range"
            else:
                ppa = base_ppa * 1.25
                status = "target_above_range"

        for i in range(max_iter):
            irr_cur = calc_irr(ppa)
            err = irr_cur - target_irr
            iterations.append({
                "iter": i,
                "ppa": round(ppa, 4),
                "irr_pct": round(irr_cur * 100, 4),
                "error_pct": round(err * 100, 4),
                "status": "converged" if abs(err) < tol else "iterating"
            })

            if abs(err) < tol:
                solution = {
                    "ppa": round(ppa, 4),
                    "irr_pct": round(irr_cur * 100, 4),
                    "error_pct": round(err * 100, 4),
                    "iterations": i + 1,
                    "converged": True
                }
                status = "converged"
                break

            # 미분 (central difference)
            irr_plus = calc_irr(ppa + h)
            irr_minus = calc_irr(ppa - h)
            derivative = (irr_plus - irr_minus) / (2 * h)

            if abs(derivative) < 1e-8:
                status = "flat_derivative"
                break

            # Newton step + 안전장치 (최대 20% 한 스텝)
            delta = -err / derivative
            max_step = base_ppa * 0.20
            if abs(delta) > max_step:
                delta = max_step if delta > 0 else -max_step
            ppa = ppa + delta

            # 음수/비정상 방지
            if ppa < 1.0:
                ppa = 1.0
            elif ppa > base_ppa * 3:
                ppa = base_ppa * 3

        if not solution:
            # 수렴 실패 - 마지막 iteration 값 사용
            last = iterations[-1] if iterations else {"ppa": base_ppa, "irr_pct": 0}
            solution = {
                "ppa": last["ppa"],
                "irr_pct": last["irr_pct"],
                "error_pct": last.get("error_pct", 0),
                "iterations": len(iterations),
                "converged": False
            }
            if status == "not_started":
                status = "max_iter_reached"

        # 추가 meta: Dev Margin 고정값 (참고용)
        base_res = _calc_engine(base_inputs)
        dev_margin_k = base_res.get("dev_margin", 0)  # $k

        return {
            "ok": True,
            "base_ppa": round(base_ppa, 2),
            "target_irr_pct": round(target_irr_pct, 2),
            "sensitivity": sensitivity,
            "iterations": iterations,
            "solution": solution,
            "status": status,
            "dev_margin_k": round(dev_margin_k, 0),
            "tolerance_pct": tol * 100,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Break-even calculation error: {str(e)}")


@app.get("/valuation/calculate/defaults")
def get_calc_defaults(user=Depends(get_current_user)):
    """Return default input values for the calculator"""
    return {
        "pv_mwac": 199, "dc_ac_ratio": 1.34,
        "bess_mw": 199, "bess_mwh": 796,
        "cf_pct": 21.24, "life": 35,
        "module_cwp": 31.5, "bos_cwp": 42.9, "ess_per_kwh": 234.5,
        "epc_cont_pct": 8.0, "owner_pct": 3.0, "softcost_pct": 5.0,
        "intercon_m": 120.0, "dev_cost_m": 20.0,
        "dev_margin_kwac": 200, "epc_margin_pct": 7.95,
        "ppa_price": 68.82, "ppa_term": 25, "ppa_esc": 0,
        "bess_toll": 13.68, "bess_toll_term": 20, "merchant_ppa": 45.0,
        "degradation": 0.0064,
        "pv_om": 4.5, "bess_om": 8.64, "insurance_pv": 10.57,
        "insurance_bess": 5.05, "asset_mgmt": 210,
        "prop_tax_yr1": 3162, "land_rent_yr1": 437, "opex_esc": 2.0,
        "aug_price": 150, "aug_mwh_pct": 18.8, "aug_y1": 4, "aug_y2": 8, "aug_y3": 12,
        "debt_ratio": 47.6, "int_rate": 5.5, "loan_term": 18,
        "te_pct": 0,
    }

# ══════════════════════════════════════════════════
#  헬스체크
# ══════════════════════════════════════════════════
@app.get("/")
def root():
    return {"status": "ok", "service": "HWR Dashboard API"}

@app.get("/health")
def health():
    return {"status": "ok", "ts": datetime.datetime.now().isoformat()}

# ══════════════════════════════════════════════════
#  프로젝트 데이터 (인허가/예산/일정/메모)
# ══════════════════════════════════════════════════
@app.get("/project/{project_id}")
def get_project(project_id: str, user=Depends(get_current_user)):
    return fb_read(f"projects/{project_id}")

@app.post("/project/{project_id}")
def save_project(project_id: str, data: dict, user=Depends(get_current_user)):
    data["updatedAt"] = datetime.datetime.now().isoformat()[:16]
    data["updatedBy"] = user["email"]
    fb_write(f"projects/{project_id}", data)
    return {"ok": True}

@app.get("/projects")
def get_all_projects(user=Depends(get_current_user)):
    return fb_read("projects")

# ══════════════════════════════════════════════════
#  Valuation (PF 모델 업로드 / 조회)
# ══════════════════════════════════════════════════

def parse_pf_model(filepath: str) -> dict:
    """xlsb에서 핵심 가정값과 아웃풋 추출"""
    assumptions = {}
    outputs = {}

    with open_workbook(filepath) as wb:

        # ── PF Intake → assumptions ───────────────
        try:
            with wb.get_sheet("PF Intake") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row if c.v is not None]
                    if len(vals) < 2:
                        continue
                    label = str(vals[0]).strip()
                    val   = vals[1] if len(vals) > 1 else None

                    intake_map = {
                        "Project Name":             ("project_name", str),
                        "PJ Characteristic":        ("technology",   str),
                        "State":                    ("state",        str),
                        "PV : Project Size (MWac)": ("pv_mwac",      float),
                        "NTP Date":                 ("ntp",          str),
                        "COD":                      ("cod",          str),
                        "DC/AC Ratio":              ("dc_ac_ratio",  float),
                        "Total Site Area":          ("site_area_ac", float),
                    }
                    if label in intake_map:
                        key, typ = intake_map[label]
                        try:
                            assumptions[key] = typ(val) if val is not None else None
                        except Exception:
                            assumptions[key] = str(val) if val is not None else None

                    # BESS: ESS size 행은 "4hr" 같은 duration값 → bess_duration으로 저장
                    if label == "ESS size (MW)":
                        assumptions["bess_duration"] = str(val) if val is not None else None

                    # BESS: ESS Duration 행에서 실제 MW 추출
                    if label == "ESS Duration (Hours)":
                        try:
                            assumptions["bess_mw"] = float(val)
                        except Exception:
                            pass

                    # BESS: ESS storage MWh — hex/int 모두 처리
                    if label == "ESS storage size (MWh)":
                        try:
                            v = val
                            if isinstance(v, str) and v.startswith("0x"):
                                assumptions["bess_mwh"] = float(int(v, 16))
                            else:
                                assumptions["bess_mwh"] = float(v)
                        except Exception:
                            assumptions["bess_mwh"] = None
        except Exception:
            pass

        # ── Quarterly Assumptions → 운영 가정 ─────
        try:
            with wb.get_sheet("Quarterly Assumptions") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row if c.v is not None]
                    if len(vals) < 3:
                        continue
                    label = str(vals[0]).strip()
                    val   = vals[2] if len(vals) > 2 else None

                    qa_map = {
                        "Degradation":              ("degradation",       float),
                        "Availability (yr 1)":      ("availability_yr1",  float),
                        "Availability (yr 2+)":     ("availability_yr2",  float),
                        "PV Covered O&M":           ("pv_om_covered",     float),
                        "PV Non-covered O&M":       ("pv_om_noncovered",  float),
                        "Asset management < 200MW": ("asset_mgmt_sm",     float),
                        "Asset management > 200MW": ("asset_mgmt_lg",     float),
                        "PV Merchant Haircut":      ("merchant_haircut",  float),
                    }
                    if label in qa_map:
                        key, typ = qa_map[label]
                        try:
                            assumptions[key] = typ(val) if val is not None else None
                        except Exception:
                            pass
        except Exception:
            pass

        # ── Summary → outputs (Case 2 = PV+BESS 컬럼 기준, index 3) ──
        try:
            with wb.get_sheet("Summary") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row if c.v is not None]
                    if len(vals) < 3:
                        continue
                    label = str(vals[0]).strip()

                    summary_map = {
                        "levered project IRR (full life)":   "levered_irr",
                        "Unlevered project IRR (full life)": "unlevered_irr",
                        "Sponsor levered IRR (full life)":   "sponsor_irr",
                        "Sponsor levered IRR (contract)":    "sponsor_irr_contract",
                        # 신규: After-Tax IRR (Class B 관점) + WACC
                        "Sponsor levered after-tax IRR (before NOL)":  "sponsor_irr_aftertax_before_nol",
                        "Sponsor levered after-tax IRR (after NOL)":   "sponsor_irr_aftertax_after_nol",
                        "Weighted average cost of capital":  "wacc",
                        "WACC":                              "wacc",
                        "Total Project Cost":                "capex_total",
                        "Debt":                              "debt",
                        "Tax Equity Investment":             "tax_equity",
                        "Sponsor Equity Investment":         "sponsor_equity",
                        "PV : PPA Price":                    "ppa_price",
                        "PV : PPA term":                     "ppa_term",
                        "BESS : Toll rate":                  "bess_toll",
                        "BESS : Toll term":                  "bess_toll_term",
                        "HQC DEV Margin (000$)":             "dev_margin",
                        "Total Margin (000$)":               "total_margin",
                    }
                    if label in summary_map:
                        key = summary_map[label]
                        # Case 2 (PV+BESS) = vals[3], fallback to vals[2]
                        val = vals[3] if len(vals) > 3 else (vals[2] if len(vals) > 2 else None)
                        try:
                            v = float(val)
                            if key in ("levered_irr", "unlevered_irr",
                                       "sponsor_irr", "sponsor_irr_contract",
                                       "sponsor_irr_aftertax_before_nol",
                                       "sponsor_irr_aftertax_after_nol",
                                       "wacc"):
                                outputs[key] = round(v, 6)
                            else:
                                outputs[key] = round(v, 2)
                        except Exception:
                            pass
        except Exception:
            pass

        # ── Returns 시트 → After-Tax IRR (Before/After NOL) 및 기타 세분화 IRR ──
        # Returns 시트의 'Sponsor net aftertax cashflow' 줄에 IRR이 있음
        # 각 IRR 값은 보통 4~5번째 컬럼 위치에 있고, 레이블은 맨 앞
        try:
            with wb.get_sheet("Returns") as ws:
                rows_list = list(ws.rows())
                # 라인 순서대로 처리 (NOL 이전 aftertax는 첫 번째 매칭, 이후는 두 번째)
                aftertax_matches = []
                unlevered_aftertax_matches = []
                for row in rows_list:
                    vals = [c.v for c in row]
                    label = ""
                    # 첫 번째 문자열 셀을 레이블로
                    for v in vals[:3]:
                        if isinstance(v, str) and v.strip():
                            label = v.strip()
                            break
                    if not label:
                        continue
                    # IRR 숫자 찾기 (0 < v < 1 범위의 float)
                    irr_val = None
                    for v in vals:
                        if isinstance(v, float) and 0.001 < v < 0.5 and v != 1.0:
                            # 첫 번째 그럴듯한 IRR 값 (label 이후)
                            irr_val = round(v, 6)
                            break
                    if irr_val is None:
                        continue

                    # Sponsor net pretax cashflow — Levered Pre-Tax
                    #   "(without ITC or PTC)" = baseline (Line 25, ~10.02%)
                    #   "(with PTC)" 는 PTC 모델이므로 제외
                    if (label.startswith("Sponsor net pretax cashflow")
                        and "unlevered" not in label.lower()
                        and "with ptc" not in label.lower()
                        and "with itc" not in label.lower()):
                        if "sponsor_irr_levered_pretax" not in outputs:
                            outputs["sponsor_irr_levered_pretax"] = irr_val
                    # Sponsor net unlevered pretax cashflow
                    elif (label.startswith("Sponsor net unlevered pretax")
                          and "with ptc" not in label.lower()
                          and "with itc" not in label.lower()):
                        if "sponsor_irr_unlevered_pretax" not in outputs:
                            outputs["sponsor_irr_unlevered_pretax"] = irr_val
                    # Sponsor net aftertax cashflow (level IRR, NOL 전/후 두 줄)
                    #   - 첫 등장 = Before NOL (~13.62%)
                    #   - 두 번째 등장 (NOL effect 처리 후) = After NOL (~10.51%)
                    # "(including Residual Value)" 및 State Tax 버전은 제외
                    elif (label == "Sponsor net aftertax cashflow"
                          and "residual" not in label.lower()
                          and "state" not in label.lower()):
                        aftertax_matches.append(irr_val)
                    # Sponsor net unlevered aftertax cashflow
                    elif (label == "Sponsor net unlevered aftertax cashflow"
                          or label == "Sponsor net unlevered aftertax cashflow with NOL"):
                        unlevered_aftertax_matches.append(irr_val)

                # 매칭 순서 기반: first = before NOL, second = after NOL
                if len(aftertax_matches) >= 1:
                    outputs["sponsor_irr_aftertax_before_nol"] = aftertax_matches[0]
                if len(aftertax_matches) >= 2:
                    # 두 번째 매칭이 After NOL (세 번째 이상은 State Tax 변형)
                    outputs["sponsor_irr_aftertax_after_nol"] = aftertax_matches[1]
                if len(unlevered_aftertax_matches) >= 1:
                    outputs["sponsor_irr_unlevered_aftertax_before_nol"] = unlevered_aftertax_matches[0]
                if len(unlevered_aftertax_matches) >= 2:
                    outputs["sponsor_irr_unlevered_aftertax_after_nol"] = unlevered_aftertax_matches[1]
        except Exception:
            pass

        # ── Sensitivities 시트 → WACC, Cost of Debt ──
        # "Weighted average cost of capital" 레이블이 있는 행에서 2번째 컬럼 값
        try:
            with wb.get_sheet("Sensitivities") as ws:
                for row in ws.rows():
                    vals = [c.v for c in row]
                    label = ""
                    for v in vals[:4]:
                        if isinstance(v, str) and v.strip():
                            label = v.strip()
                            break
                    if not label:
                        continue
                    low = label.lower()
                    # WACC - "Weighted average cost of capital"
                    if "weighted average cost of capital" in low and "wacc" not in outputs:
                        for v in vals:
                            if isinstance(v, float) and 0.01 < v < 0.3:
                                outputs["wacc"] = round(v, 6)
                                break
                    # Cost of debt
                    elif label.lower().strip() == "cost of debt" and "cost_of_debt" not in outputs:
                        for v in vals:
                            if isinstance(v, float) and 0.01 < v < 0.3:
                                outputs["cost_of_debt"] = round(v, 6)
                                break
        except Exception:
            pass

    # ── CF_Annual → 연도별 실제 수익 추출
    try:
        with wb.get_sheet("CF_Annual") as ws:
            for row in ws.rows():
                vals = [c.v for c in row if c.v is not None]
                if len(vals) < 5: continue
                label = str(vals[0]).strip()
                # Y1 시작 인덱스: 앞 4개(total, pre-COD, 0, 0) 제거 후 운영연도
                op_vals = [v for v in vals[1:] if isinstance(v, (int, float))]

                if "PPA #2 BESS" in label and "Revenue" in label:
                    try:
                        bess_rev_y1 = float(op_vals[4]) if len(op_vals) > 4 else 0
                        bess_mw = assumptions.get("bess_mw") or 199
                        if bess_rev_y1 > 0 and bess_mw > 0:
                            outputs["bess_toll_y1_effective"] = round(bess_rev_y1/(bess_mw*1000*12)*1000, 4)
                            outputs["bess_rev_y1"] = round(bess_rev_y1, 0)
                        # 연도별 BESS 수익 (인덱스 4~38 = Y1~Y35)
                        outputs["bess_rev_schedule"] = [round(float(v),0) for v in op_vals[4:39] if isinstance(v,(int,float))]
                    except: pass

                if "PPA #1 PV" in label and "Revenue" in label:
                    try:
                        pv_rev_y1 = float(op_vals[4]) if len(op_vals) > 4 else 0
                        if pv_rev_y1 > 0:
                            outputs["pv_rev_y1"] = round(pv_rev_y1, 0)
                        outputs["pv_rev_schedule"] = [round(float(v),0) for v in op_vals[4:39] if isinstance(v,(int,float))]
                    except: pass

                if "Merchant PV Power Revenue" in label:
                    try:
                        outputs["merch_rev_schedule"] = [round(float(v),0) for v in op_vals[4:39] if isinstance(v,(int,float))]
                    except: pass
    except Exception:
        pass

    # bess_mwh 보정: xlsb hex 파싱 한계 → pv_mwac × duration(숫자)으로 계산
    try:
        duration_str = assumptions.get("bess_duration", "")
        duration_h = float("".join(x for x in str(duration_str) if x.isdigit() or x=="."))
        pv_mwac = assumptions.get("pv_mwac") or assumptions.get("bess_mw")
        if pv_mwac and duration_h:
            assumptions["bess_mwh"] = round(float(pv_mwac) * duration_h, 1)
    except Exception:
        pass

    return {"assumptions": assumptions, "outputs": outputs}


@app.post("/valuation/upload")
async def upload_valuation(
    project_id: str = Form(...),
    scenario:   str = Form(default=""),
    reason:     str = Form(default=""),
    approver:   str = Form(default=""),
    file: UploadFile = File(...),
    user=Depends(get_current_user)
):
    """PF 재무모델(xlsb/xlsx) 업로드 → AI 파싱 → Firebase 저장"""
    if not (file.filename.endswith(".xlsb") or file.filename.endswith(".xlsx")):
        raise HTTPException(400, "xlsb 또는 xlsx 파일만 업로드 가능합니다.")

    suffix = ".xlsb" if file.filename.endswith(".xlsb") else ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        parsed = parse_pf_model(tmp_path)
    except Exception as e:
        raise HTTPException(500, f"모델 파싱 실패: {str(e)}")
    finally:
        os.unlink(tmp_path)

    ts      = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    safe_id = project_id.replace("/", "_").replace(".", "_")

    payload = {
        "uploaded_at": datetime.datetime.now().isoformat(),
        "uploaded_by": user["email"],
        "filename":    file.filename,
        "scenario":    scenario,
        "reason":      reason,
        "approver":    approver,
        "assumptions": parsed["assumptions"],
        "outputs":     parsed["outputs"],
    }

    fb_put(f"valuation/{safe_id}/versions/{ts}", payload)
    fb_put(f"valuation/{safe_id}/latest", payload)

    return {
        "ok":         True,
        "project_id": safe_id,
        "timestamp":  ts,
        "parsed":     parsed,
    }


@app.get("/valuation")
def get_all_valuations(user=Depends(get_current_user)):
    """전체 프로젝트 latest 비교 (Valuation 탭용)"""
    all_data = fb_read("valuation") or {}
    result = {}
    for pid, pdata in all_data.items():
        if isinstance(pdata, dict) and "latest" in pdata:
            result[pid] = pdata["latest"]
    return result


@app.get("/valuation/{project_id}")
def get_valuation(project_id: str, user=Depends(get_current_user)):
    safe_id = project_id.replace("/", "_").replace(".", "_")
    return fb_read(f"valuation/{safe_id}")


@app.get("/valuation/{project_id}/latest")
def get_valuation_latest(project_id: str, user=Depends(get_current_user)):
    safe_id = project_id.replace("/", "_").replace(".", "_")
    return fb_read(f"valuation/{safe_id}/latest")


@app.get("/valuation/{project_id}/versions")
def get_valuation_versions(project_id: str, user=Depends(get_current_user)):
    safe_id = project_id.replace("/", "_").replace(".", "_")
    return fb_read(f"valuation/{safe_id}/versions")


@app.post("/valuation/generate-ic-summary")
async def generate_ic_summary(payload: dict, user=Depends(get_current_user)):
    """Claude API로 IC Summary 전문 보고서 생성"""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    proj    = payload.get("project_name", "Project")
    metrics = payload.get("metrics", {})
    scenarios = payload.get("scenarios", [])
    assumptions = payload.get("assumptions", {})
    history = payload.get("history", [])
    today   = payload.get("date", "")

    scen_text = ""
    if scenarios:
        scen_text = "\n\nScenario Analysis:\n"
        for s in scenarios:
            scen_text += f"  {s.get('name','')}: IRR {s.get('irr','—')}, Dev Margin {s.get('margin','—')}\n"

    hist_text = ""
    if history:
        hist_text = "\n\nVersion History (recent):\n"
        for h in history[:3]:
            hist_text += f"  {h.get('date','')} — {h.get('reason','')}\n"

    prompt = (
        "You are a senior investment analyst at a US renewable energy developer. "
        "Write a concise, professional Investment Committee (IC) Summary in Korean (with key financial metrics in English). "
        "Use formal Korean business writing style. Structure it with clear sections.\n\n"
        f"Project: {proj}\n"
        f"Date: {today}\n\n"
        "Financial Metrics:\n"
        f"  Sponsor IRR: {metrics.get('sirr','—')}\n"
        f"  Dev Margin: {metrics.get('dev_margin','—')}\n"
        f"  Levered IRR: {metrics.get('lirr','—')}\n"
        f"  Unlevered IRR: {metrics.get('uirr','—')}\n"
        f"  EBITDA Yield: {metrics.get('ebitda_yield','—')}\n"
        f"  Total CAPEX: {metrics.get('capex','—')}\n"
        f"  Debt: {metrics.get('debt','—')} ({metrics.get('debt_pct','—')})\n"
        f"  Tax Equity: {metrics.get('te','—')}\n"
        f"  Sponsor Equity: {metrics.get('eq','—')}\n"
        f"  PPA: {metrics.get('ppa','—')}\n"
        f"  BESS Toll: {metrics.get('toll','—')}\n"
        f"  ITC/PTC: {metrics.get('credit','—')}\n"
        f"  Flip Yield: {metrics.get('flip','—')}\n"
        f"{scen_text}{hist_text}\n\n"
        "Write the IC Summary with these sections:\n"
        "1. 프로젝트 개요 (2-3 sentences)\n"
        "2. 핵심 재무 지표 (bullet points with brief commentary)\n"
        "3. Deal Structure 특징 (TE flip structure, debt terms 등)\n"
        "4. 리스크 요인 (2-3 key risks)\n"
        "5. 투자 의견 (1 paragraph recommendation)\n\n"
        "Keep it concise — suitable for a 1-page print. "
        "Return ONLY valid JSON: "
        '{{"sections":[{{"title":"섹션제목","content":"내용"}}]}}'
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=40
    )
    if resp.status_code != 200:
        raise HTTPException(500, f"Claude API 오류: {resp.text[:200]}")

    data = resp.json()
    text = "".join(b.get("text","") for b in data.get("content",[]))
    clean = text.replace("```json","").replace("```","").strip()
    return {"ok": True, "result": clean}


# ══════════════════════════════════════════════════
#  IC Summary PDF Export (WeasyPrint — world-class formatting)
# ══════════════════════════════════════════════════
import base64 as _base64
from fastapi.responses import Response as _Response

def _esc_html(s):
    """HTML escape helper."""
    if s is None:
        return ""
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))

def _fmt_pct(v, decimals=2):
    """숫자 → 퍼센트 문자열. 이미 문자열이면 그대로."""
    if v is None or v == "—":
        return "—"
    if isinstance(v, str):
        return v
    try:
        return f"{float(v)*100:.{decimals}f}%"
    except Exception:
        return "—"

def _fmt_usd_m(v):
    """$M 포맷 (input in thousands)."""
    if v is None or v == "—":
        return "—"
    if isinstance(v, str):
        return v
    try:
        return f"${float(v)/1000:.1f}M"
    except Exception:
        return "—"

def _build_ic_pdf_html(data: dict) -> str:
    """IC Summary HTML (WeasyPrint용). World-class IB/PE 수준 포맷."""
    proj_name = data.get("project_name", "Project")
    today = data.get("date", datetime.date.today().isoformat())
    verdict = (data.get("verdict") or "").upper() or "—"
    verdict_color = data.get("verdict_color", "amber")

    # 색상 매핑
    color_map = {
        "green": "#059669",  # emerald-600
        "amber": "#D97706",  # amber-600
        "red":   "#DC2626",  # red-600
    }
    v_color = color_map.get(verdict_color, "#6B7280")

    # 지표
    outputs = data.get("outputs", {}) or {}
    assumptions = data.get("assumptions", {}) or {}
    pv_mwac = assumptions.get("pv_mwac") or outputs.get("pv_mwac") or "—"
    bess_mw = assumptions.get("bess_mw") or "—"
    cod = assumptions.get("cod") or "—"
    ntp = assumptions.get("ntp") or "—"
    state = data.get("state") or assumptions.get("state") or "—"
    iso = data.get("iso") or assumptions.get("iso") or "—"

    # 5 IRR 지표
    irr_lev_pre  = _fmt_pct(outputs.get("sponsor_irr_levered_pretax") or outputs.get("sponsor_irr"))
    irr_at_before = _fmt_pct(outputs.get("sponsor_irr_aftertax_before_nol"))
    irr_at_after  = _fmt_pct(outputs.get("sponsor_irr_aftertax_after_nol"))
    irr_unlev    = _fmt_pct(outputs.get("sponsor_irr_unlevered_pretax") or outputs.get("unlevered_irr"))
    wacc_val     = _fmt_pct(outputs.get("wacc"))

    # 재무 요약
    capex = _fmt_usd_m(outputs.get("capex_total"))
    debt  = _fmt_usd_m(outputs.get("debt"))
    te    = _fmt_usd_m(outputs.get("tax_equity"))
    eq    = _fmt_usd_m(outputs.get("sponsor_equity"))
    dev_margin = _fmt_usd_m(outputs.get("dev_margin"))
    margin_cwp = outputs.get("margin_cwp")
    margin_cwp_str = f"{margin_cwp:.2f} c/Wp" if isinstance(margin_cwp, (int, float)) else "—"
    ppa_price = outputs.get("ppa_price") or assumptions.get("ppa_price") or "—"
    ppa_term  = outputs.get("ppa_term") or assumptions.get("ppa_term") or "—"
    bess_toll = outputs.get("bess_toll") or assumptions.get("bess_toll") or "—"
    ebitda_y  = outputs.get("ebitda_yield")
    ebitda_y_str = f"{ebitda_y:.2f}%" if isinstance(ebitda_y, (int, float)) else "—"

    # AI 분석 결과 (IC Opinion 에서 생성된 것)
    ic_analysis = data.get("ic_analysis", {}) or {}
    thesis = ic_analysis.get("thesis", "")
    rec    = ic_analysis.get("rec", "")
    risks  = ic_analysis.get("risks", []) or []
    threshold_status = ic_analysis.get("threshold_status", {}) or {}
    dev_ic = ic_analysis.get("dev_ic", {}) or {}

    # Sensitivity (프론트에서 계산된 값)
    scenarios = data.get("scenarios", []) or []

    # Threshold 메타
    thresholds = data.get("thresholds", {}) or {}
    thr_irr = thresholds.get("sponsor_irr_pct", 9.0)
    thr_margin = thresholds.get("dev_margin_cwp", 10.0)

    # 리스크 분리: compliance_count만큼 앞쪽은 고정 체크리스트, 뒤는 AI 생성
    compliance_count = int(ic_analysis.get("compliance_count", 0) or 0)
    compliance_items = risks[:compliance_count] if compliance_count else []
    ai_risks = risks[compliance_count:] if compliance_count else risks

    # 1) 컴플라이언스 체크리스트 HTML
    compliance_html = ""
    for r in compliance_items:
        title = _esc_html(r.get("title", ""))
        detail = _esc_html(r.get("detail", ""))
        sev = r.get("severity", "Watch")
        compliance_html += f"""
        <div class="compliance-item">
          <div class="compliance-box"></div>
          <div class="compliance-body">
            <div class="compliance-head">
              <span class="compliance-title">{title}</span>
              <span class="compliance-sev">{sev}</span>
            </div>
            <div class="compliance-detail">{detail}</div>
          </div>
        </div>
        """

    # 2) AI 프로젝트별 리스크 HTML
    risks_html = ""
    sev_color = {"Critical": "#DC2626", "Watch": "#D97706", "OK": "#059669"}
    for i, r in enumerate(ai_risks[:8]):
        sev = r.get("severity", "OK")
        c = sev_color.get(sev, "#6B7280")
        title = _esc_html(r.get("title", ""))
        detail = _esc_html(r.get("detail", ""))
        risks_html += f"""
        <div class="risk-item">
          <div class="risk-header">
            <span class="risk-num">{i+1:02d}</span>
            <span class="risk-title">{title}</span>
            <span class="risk-sev" style="background:{c}">{sev}</span>
          </div>
          <div class="risk-detail">{detail}</div>
        </div>
        """

    # Scenario 테이블
    scen_rows = ""
    for s in scenarios:
        scen_rows += f"""
        <tr>
          <td class="scen-name">{_esc_html(s.get('name','—'))}</td>
          <td class="scen-val">{_esc_html(s.get('irr','—'))}</td>
          <td class="scen-val">{_esc_html(s.get('margin','—'))}</td>
        </tr>
        """

    # Threshold 체크
    def _chk(ok):
        return ('<span style="color:#059669;font-weight:700">✓ PASS</span>' if ok
                else '<span style="color:#DC2626;font-weight:700">✗ FAIL</span>')
    thr_irr_ok = threshold_status.get("irr_ok", False)
    thr_margin_ok = threshold_status.get("margin_ok", False)
    thr_irr_gap = _esc_html(threshold_status.get("irr_gap", ""))
    thr_margin_gap = _esc_html(threshold_status.get("margin_gap", ""))

    # HTML 조립
    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>IC Summary - {_esc_html(proj_name)}</title>
<style>
@page {{
  size: A4 portrait;
  margin: 18mm 16mm 18mm 16mm;
  @bottom-center {{
    content: counter(page) " / " counter(pages);
    font-family: 'Noto Sans KR', 'Helvetica', sans-serif;
    font-size: 8pt;
    color: #6B7280;
  }}
  @bottom-left {{
    content: "Hanwha Energy USA Holdings · Internal IC Memo";
    font-family: 'Noto Sans KR', 'Helvetica', sans-serif;
    font-size: 7pt;
    color: #9CA3AF;
  }}
  @bottom-right {{
    content: "{_esc_html(today)}";
    font-family: 'Noto Sans KR', 'Helvetica', sans-serif;
    font-size: 7pt;
    color: #9CA3AF;
  }}
}}
@page :first {{
  @bottom-center {{ content: none; }}
  @bottom-left {{ content: none; }}
  @bottom-right {{ content: none; }}
}}
* {{ box-sizing: border-box; }}
body {{
  font-family: 'Noto Sans KR', 'Helvetica Neue', Helvetica, sans-serif;
  font-size: 10pt;
  line-height: 1.55;
  color: #111827;
  margin: 0;
  padding: 0;
  -webkit-font-smoothing: antialiased;
}}

/* ── Cover ────────────────────────────────────── */
.cover {{
  height: 260mm;
  display: flex;
  flex-direction: column;
  justify-content: space-between;
  padding: 20mm 8mm 8mm 8mm;
}}
.cover-header {{
  font-size: 8pt;
  font-weight: 600;
  color: #6B7280;
  letter-spacing: 0.2em;
  text-transform: uppercase;
  border-bottom: 1px solid #E5E7EB;
  padding-bottom: 12pt;
}}
.cover-main {{
  margin-top: 40mm;
}}
.cover-tag {{
  font-size: 9pt;
  font-weight: 600;
  color: {v_color};
  letter-spacing: 0.16em;
  text-transform: uppercase;
  margin-bottom: 10pt;
}}
.cover-title {{
  font-size: 36pt;
  font-weight: 800;
  color: #111827;
  letter-spacing: -1.2pt;
  line-height: 1.05;
  margin-bottom: 14pt;
}}
.cover-sub {{
  font-size: 12pt;
  color: #4B5563;
  font-weight: 400;
  margin-bottom: 40pt;
}}
.cover-verdict {{
  display: inline-block;
  padding: 10pt 22pt;
  border: 2pt solid {v_color};
  border-radius: 2pt;
  font-size: 22pt;
  font-weight: 800;
  color: {v_color};
  letter-spacing: 4pt;
}}
.cover-stats {{
  display: flex;
  gap: 20pt;
  margin-top: 24pt;
}}
.cover-stat {{
  flex: 1;
  border-left: 2pt solid #E5E7EB;
  padding-left: 10pt;
}}
.cover-stat-label {{
  font-size: 7pt;
  font-weight: 700;
  color: #6B7280;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  margin-bottom: 3pt;
}}
.cover-stat-value {{
  font-size: 16pt;
  font-weight: 700;
  color: #111827;
  font-variant-numeric: tabular-nums;
}}
.cover-footer {{
  margin-top: auto;
  padding-top: 20pt;
  border-top: 1px solid #E5E7EB;
  display: flex;
  justify-content: space-between;
  font-size: 8pt;
  color: #6B7280;
}}

/* ── Content Pages ────────────────────────────── */
.page-break {{ page-break-before: always; }}

h1 {{
  font-size: 14pt;
  font-weight: 800;
  color: #111827;
  margin: 0 0 3pt 0;
  letter-spacing: -0.3pt;
}}
.section-sub {{
  font-size: 8pt;
  color: #6B7280;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  margin-bottom: 12pt;
  border-bottom: 1pt solid #E5E7EB;
  padding-bottom: 6pt;
}}
h2 {{
  font-size: 10pt;
  font-weight: 700;
  color: #111827;
  margin: 14pt 0 6pt 0;
  letter-spacing: 0;
}}
p {{ margin: 4pt 0; color: #1F2937; }}

/* Metrics Grid */
.metrics-grid {{
  display: grid;
  grid-template-columns: repeat(5, 1fr);
  gap: 4pt;
  margin-bottom: 14pt;
}}
.metric-card {{
  padding: 7pt 8pt;
  border: 0.5pt solid #D1D5DB;
  border-radius: 2pt;
  overflow: hidden;
}}
.metric-card-primary {{
  border-left: 2.5pt solid #059669;
}}
.metric-card-secondary {{
  border-left: 2.5pt solid #D97706;
}}
.metric-card-wacc {{
  border-left: 2.5pt solid #2563EB;
}}
.metric-label {{
  font-size: 6.5pt;
  font-weight: 700;
  color: #6B7280;
  letter-spacing: 0.04em;
  text-transform: uppercase;
  margin-bottom: 2pt;
  white-space: nowrap;
}}
.metric-value {{
  font-size: 14pt;
  font-weight: 700;
  color: #111827;
  font-variant-numeric: tabular-nums;
  line-height: 1.1;
  white-space: nowrap;
}}
.metric-sub {{
  font-size: 6.5pt;
  color: #6B7280;
  margin-top: 2pt;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}}

/* Financial table */
.fin-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 9pt;
  margin: 8pt 0 14pt 0;
}}
.fin-table th {{
  text-align: left;
  padding: 6pt 8pt;
  border-bottom: 1pt solid #111827;
  font-size: 7pt;
  font-weight: 700;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  color: #374151;
}}
.fin-table td {{
  padding: 5pt 8pt;
  border-bottom: 0.3pt solid #E5E7EB;
  font-variant-numeric: tabular-nums;
}}
.fin-table td.val {{ text-align: right; font-weight: 600; }}
.fin-table tr.subtotal td {{
  background: #F9FAFB;
  font-weight: 700;
  border-top: 0.5pt solid #6B7280;
}}

/* Threshold check */
.thr-box {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10pt;
  margin: 10pt 0;
}}
.thr-item {{
  padding: 10pt 12pt;
  border: 1pt solid #E5E7EB;
  border-radius: 2pt;
  background: #F9FAFB;
}}
.thr-label {{
  font-size: 8pt;
  font-weight: 600;
  color: #6B7280;
  margin-bottom: 4pt;
}}
.thr-status {{ font-size: 11pt; margin-bottom: 3pt; }}
.thr-gap {{ font-size: 9pt; color: #374151; }}

/* Thesis / Recommendation boxes */
.thesis-box {{
  padding: 12pt 14pt;
  background: #F9FAFB;
  border-left: 3pt solid #2563EB;
  border-radius: 0 2pt 2pt 0;
  margin: 10pt 0;
  font-size: 10pt;
  line-height: 1.7;
  color: #1F2937;
}}
.rec-box {{
  padding: 14pt 16pt;
  background: #FEF3C7;
  border-left: 3pt solid #D97706;
  border-radius: 0 2pt 2pt 0;
  margin: 10pt 0;
  font-size: 10pt;
  line-height: 1.7;
  color: #78350F;
  font-weight: 500;
}}

/* Risk items */
.risk-item {{
  padding: 10pt 0;
  border-bottom: 0.5pt solid #E5E7EB;
}}
.risk-item:last-child {{ border-bottom: none; }}
.risk-header {{
  display: flex;
  align-items: center;
  gap: 8pt;
  margin-bottom: 4pt;
}}
.risk-num {{
  font-size: 8pt;
  font-weight: 700;
  color: #9CA3AF;
  font-variant-numeric: tabular-nums;
  min-width: 18pt;
}}
.risk-title {{
  font-size: 10pt;
  font-weight: 700;
  color: #111827;
  flex: 1;
}}
.risk-sev {{
  color: #fff;
  font-size: 7pt;
  font-weight: 700;
  padding: 2pt 7pt;
  border-radius: 2pt;
  letter-spacing: 0.05em;
}}
.risk-detail {{
  font-size: 9pt;
  color: #4B5563;
  line-height: 1.6;
  margin-left: 26pt;
}}

/* Compliance Checklist items */
.compliance-note {{
  font-size: 8pt;
  color: #6B7280;
  font-style: italic;
  margin-bottom: 8pt;
}}
.compliance-item {{
  display: flex;
  gap: 10pt;
  padding: 9pt 12pt;
  background: #FFFBEB;
  border: 0.5pt solid #FCD34D;
  border-left: 3pt solid #D97706;
  border-radius: 2pt;
  margin-bottom: 6pt;
}}
.compliance-box {{
  flex-shrink: 0;
  width: 12pt;
  height: 12pt;
  border: 1pt solid #9CA3AF;
  border-radius: 2pt;
  margin-top: 2pt;
}}
.compliance-body {{ flex: 1; }}
.compliance-head {{
  display: flex;
  align-items: center;
  gap: 8pt;
  margin-bottom: 3pt;
}}
.compliance-title {{
  font-size: 10pt;
  font-weight: 700;
  color: #111827;
  flex: 1;
}}
.compliance-sev {{
  font-size: 7pt;
  font-weight: 700;
  color: #D97706;
  border: 0.5pt solid #D97706;
  padding: 1pt 6pt;
  border-radius: 8pt;
  letter-spacing: 0.05em;
}}
.compliance-detail {{
  font-size: 9pt;
  color: #78350F;
  line-height: 1.6;
}}

/* Scenario table */
.scen-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 9pt;
  margin: 8pt 0;
}}
.scen-table th {{
  text-align: left;
  padding: 7pt 10pt;
  background: #111827;
  color: #fff;
  font-size: 7.5pt;
  font-weight: 700;
  letter-spacing: 0.1em;
  text-transform: uppercase;
}}
.scen-table td {{
  padding: 7pt 10pt;
  border-bottom: 0.5pt solid #E5E7EB;
}}
.scen-name {{ font-weight: 700; color: #111827; }}
.scen-val {{ font-variant-numeric: tabular-nums; text-align: right; }}

/* Dev IC Grid */
.devic-grid {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10pt;
  margin: 10pt 0;
}}
.devic-item {{
  padding: 9pt 12pt;
  border: 0.5pt solid #E5E7EB;
  border-radius: 2pt;
  background: #FEFEFE;
}}
.devic-label {{
  font-size: 7pt;
  font-weight: 700;
  color: #6B7280;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  margin-bottom: 3pt;
}}
.devic-value {{
  font-size: 9pt;
  color: #1F2937;
  line-height: 1.5;
}}

/* Footer note */
.confidential-note {{
  margin-top: 20pt;
  padding-top: 10pt;
  border-top: 0.3pt solid #E5E7EB;
  font-size: 7pt;
  color: #9CA3AF;
  font-style: italic;
  text-align: center;
}}
</style>
</head>
<body>

<!-- ══ COVER PAGE ══ -->
<div class="cover">
  <div>
    <div class="cover-header">Hanwha Energy USA Holdings · Investment Committee Memo</div>
  </div>

  <div class="cover-main">
    <div class="cover-tag">Confidential · Internal Use Only</div>
    <div class="cover-title">{_esc_html(proj_name)}</div>
    <div class="cover-sub">{_esc_html(pv_mwac)} MWac Solar + BESS · {_esc_html(state)} ({_esc_html(iso)}) · COD {_esc_html(cod)}</div>

    <div class="cover-verdict">{_esc_html(verdict)}</div>

    <div class="cover-stats">
      <div class="cover-stat">
        <div class="cover-stat-label">Sponsor IRR</div>
        <div class="cover-stat-value">{irr_lev_pre}</div>
      </div>
      <div class="cover-stat">
        <div class="cover-stat-label">Dev Margin</div>
        <div class="cover-stat-value">{dev_margin}</div>
      </div>
      <div class="cover-stat">
        <div class="cover-stat-label">Total CAPEX</div>
        <div class="cover-stat-value">{capex}</div>
      </div>
      <div class="cover-stat">
        <div class="cover-stat-label">WACC</div>
        <div class="cover-stat-value">{wacc_val}</div>
      </div>
    </div>
  </div>

  <div class="cover-footer">
    <span>Prepared: {_esc_html(today)}</span>
    <span>{_esc_html(data.get("prepared_by",""))}</span>
  </div>
</div>

<!-- ══ PAGE 2 — EXECUTIVE SUMMARY ══ -->
<div class="page-break">
  <h1>Executive Summary</h1>
  <div class="section-sub">투자 의견 · 핵심 논거</div>

  <h2>투자 근거 (Investment Rationale)</h2>
  <div class="thesis-box">{_esc_html(thesis) if thesis else "(AI 분석 미완료 — IC Opinion 탭에서 Run AI Analysis 실행 후 재생성)"}</div>

  <h2>Recommendation</h2>
  <div class="rec-box">{_esc_html(rec) if rec else "(AI 분석 미완료)"}</div>
</div>

<!-- ══ PAGE 3 — FINANCIAL SUMMARY ══ -->
<div class="page-break">
  <h1>Financial Summary</h1>
  <div class="section-sub">재무 지표 · 자본 구조 · 계약 조건</div>

  <h2>Returns Detail</h2>
  <div class="metrics-grid">
    <div class="metric-card metric-card-primary">
      <div class="metric-label">Sponsor IRR</div>
      <div class="metric-value">{irr_lev_pre}</div>
      <div class="metric-sub">Lev · Pre-Tax</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Sponsor IRR</div>
      <div class="metric-value">{irr_at_before}</div>
      <div class="metric-sub">A-Tax · Pre-NOL</div>
    </div>
    <div class="metric-card metric-card-secondary">
      <div class="metric-label">Sponsor IRR</div>
      <div class="metric-value">{irr_at_after}</div>
      <div class="metric-sub">A-Tax · Post-NOL</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Project IRR</div>
      <div class="metric-value">{irr_unlev}</div>
      <div class="metric-sub">Unlev · Pre-Tax</div>
    </div>
    <div class="metric-card metric-card-wacc">
      <div class="metric-label">WACC</div>
      <div class="metric-value">{wacc_val}</div>
      <div class="metric-sub">Capital Cost</div>
    </div>
  </div>

  <h2>Investment Thresholds (기준 달성)</h2>
  <div class="thr-box">
    <div class="thr-item">
      <div class="thr-label">Sponsor IRR (After-TE-Flip, Full Life · min {thr_irr}%)</div>
      <div class="thr-status">{_chk(thr_irr_ok)}</div>
      <div class="thr-gap">{thr_irr_gap}</div>
    </div>
    <div class="thr-item">
      <div class="thr-label">Dev Margin (min {thr_margin} c/Wp)</div>
      <div class="thr-status">{_chk(thr_margin_ok)}</div>
      <div class="thr-gap">{thr_margin_gap}</div>
    </div>
  </div>

  <h2>Capital Structure & Deal Terms</h2>
  <table class="fin-table">
    <thead><tr><th>Item</th><th style="text-align:right">Value</th></tr></thead>
    <tbody>
      <tr><td>Total CAPEX</td><td class="val">{capex}</td></tr>
      <tr><td>Senior Debt</td><td class="val">{debt}</td></tr>
      <tr><td>Tax Equity</td><td class="val">{te}</td></tr>
      <tr><td>Sponsor Equity</td><td class="val">{eq}</td></tr>
      <tr class="subtotal"><td>Dev Margin</td><td class="val">{dev_margin} ({margin_cwp_str})</td></tr>
      <tr><td>EBITDA Yield (Y1)</td><td class="val">{ebitda_y_str}</td></tr>
      <tr><td>PPA Price × Term</td><td class="val">${_esc_html(ppa_price)}/MWh × {_esc_html(ppa_term)}yr</td></tr>
      <tr><td>BESS Toll</td><td class="val">${_esc_html(bess_toll)}/kW-mo</td></tr>
    </tbody>
  </table>

  <h2>Scenario Analysis</h2>
  <table class="scen-table">
    <thead><tr><th>Scenario</th><th style="text-align:right">Sponsor IRR</th><th style="text-align:right">Dev Margin</th></tr></thead>
    <tbody>{scen_rows if scen_rows else '<tr><td colspan="3" style="color:#9CA3AF;text-align:center">시나리오 미실행</td></tr>'}</tbody>
  </table>
</div>

<!-- ══ PAGE 4 — RISK ASSESSMENT ══ -->
<div class="page-break">
  <h1>Risk Assessment</h1>
  <div class="section-sub">규정 준수 체크 · 프로젝트별 리스크 (의사결정에 반영되지 않음)</div>

  <h2 style="margin-top:10pt">📋 규정 준수 체크리스트 (IC 승인 전 확인 필수)</h2>
  <div class="compliance-note">고정 체크리스트 · 모든 프로젝트 공통 적용</div>
  {compliance_html if compliance_html else '<p style="color:#9CA3AF;font-size:9pt">체크리스트 없음</p>'}

  <h2 style="margin-top:18pt">🔍 프로젝트별 리스크 (AI 모니터링)</h2>
  <div class="compliance-note">정보 제공 · 경제성 판정에 영향 없음</div>
  {risks_html if risks_html else '<p style="color:#9CA3AF;font-size:9pt">AI 분석 미완료 — IC Opinion 탭에서 Run AI Analysis 실행 후 재생성</p>'}

  <div class="confidential-note">
    본 문서는 Hanwha Energy USA Holdings 내부 투자심의 목적으로만 작성되었으며, 외부 유출을 금합니다.<br>
    수치 및 가정은 {_esc_html(today)} 기준 엑셀 재무모델 및 시장 데이터를 근거로 하며, 시장 변동에 따라 달라질 수 있습니다.<br>
    경제성 판정(PROCEED/RECUT/STOP)은 Dev Margin · Sponsor IRR · Unlev IRR vs WACC 기준의 순수 경제 분석 결과이며, 규정 준수 체크리스트와 개별 리스크는 별도 관리 대상입니다.
  </div>
</div>

</body>
</html>"""
    return html


@app.post("/valuation/export-pdf")
async def export_ic_pdf(payload: dict, user=Depends(get_current_user)):
    """IC Summary PDF 생성 (WeasyPrint, world-class formatting)."""
    import traceback
    import sys

    # Step 1: WeasyPrint import
    try:
        from weasyprint import HTML
        print(f"[export-pdf] WeasyPrint import OK", flush=True)
    except Exception as e:
        print(f"[export-pdf] WeasyPrint import FAILED: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"WeasyPrint import 실패: {str(e)[:300]}")

    # Step 2: HTML 문자열 생성
    try:
        html_str = _build_ic_pdf_html(payload)
        print(f"[export-pdf] HTML built, length={len(html_str)}", flush=True)
    except Exception as e:
        print(f"[export-pdf] HTML build FAILED: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"HTML 생성 오류: {str(e)[:300]}")

    # Step 3: PDF 렌더링
    try:
        pdf_bytes = HTML(string=html_str).write_pdf()
        print(f"[export-pdf] PDF rendered, size={len(pdf_bytes)} bytes", flush=True)
    except Exception as e:
        print(f"[export-pdf] PDF render FAILED: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"PDF 렌더링 오류: {str(e)[:300]}")

    proj_name = payload.get("project_name", "IC_Summary").replace(" ", "_")
    date_str = payload.get("date", datetime.date.today().isoformat()).replace("-", "")
    filename = f"IC_Summary_{proj_name}_{date_str}.pdf"

    return _Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ── WeasyPrint 진단용 미니멀 테스트 엔드포인트 ───────
@app.get("/valuation/export-pdf-test")
async def export_pdf_test(user=Depends(get_current_user)):
    """WeasyPrint가 살아있는지 간단히 테스트."""
    import traceback, sys
    try:
        from weasyprint import HTML
        simple_html = "<html><body><h1>Test</h1><p>안녕하세요, WeasyPrint 테스트</p></body></html>"
        pdf = HTML(string=simple_html).write_pdf()
        return _Response(
            content=pdf,
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="weasyprint_test.pdf"'}
        )
    except Exception as e:
        traceback.print_exc(file=sys.stdout)
        raise HTTPException(500, f"테스트 실패: {str(e)[:500]}")


@app.post("/valuation/analyze-cf")
async def analyze_cf(payload: dict, user=Depends(get_current_user)):
    """CF 데이터를 Claude API로 분석하여 인사이트 반환"""
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수 미설정")

    cf_text   = payload.get("cf_text", "")
    proj_name = payload.get("project_name", "프로젝트")

    context        = payload.get("context", "")
    proj_context   = payload.get("project_context", "")  # PPV 탭 프로젝트 메타데이터
    lang           = payload.get("lang", "en")
    mode           = payload.get("mode", "full")
    proj_meta      = payload.get("project_meta", {})
    stage    = proj_meta.get("stage", "")
    iso      = proj_meta.get("iso", "")
    proj_type= proj_meta.get("type", "")
    ntp_date = proj_meta.get("ntp", "")
    cod_date = proj_meta.get("cod", "")
    risk_pct = proj_meta.get("risk_factor", "")
    if risk_pct != "": risk_pct = f"{float(risk_pct)*100:.0f}%"
    itc_risk = proj_meta.get("itc_expiry_risk", "")
    proj_ctx = proj_meta.get("proj_ctx", "")
    thresholds     = payload.get("thresholds", {})
    current_metrics= payload.get("current_metrics", {})

    irr_thr    = thresholds.get("sponsor_irr_pct", 9.0)
    margin_thr = thresholds.get("dev_margin_cwp", 10.0)
    itc_thr    = thresholds.get("itc_min_pct", 30.0)

    curr_irr    = current_metrics.get("sponsor_irr_pct", "?")
    curr_irr_basis = current_metrics.get("sponsor_irr_basis", "After-TE-Flip, Full Life")
    curr_margin = current_metrics.get("dev_margin_cwp", "?")
    curr_itc    = current_metrics.get("itc_rate_pct", "?")
    sponsor_npv_m = current_metrics.get("sponsor_npv_m")  # $M (optional)
    project_npv_m = current_metrics.get("project_npv_m")  # $M (optional)
    ppa_term    = current_metrics.get("ppa_term", "?")
    toll_term   = current_metrics.get("toll_term", "?")
    pv_mwac     = current_metrics.get("pv_mwac", "?")

    if mode == "interp":
        prompt = (
            "미국 태양광+BESS PF 전문가로서 아래 연도별 Sponsor CF 패턴을 분석해줘.\n"
            f"프로젝트: {proj_name}\n"
            f"CF: {cf_text}\n\n"
            "3~4개 핵심 인사이트를 JSON으로 반환 (다른 텍스트 없이):\n"
            '{"insights":[{"title":"제목","detail":"설명(80자이내)"}]}'
        )
    else:
        # 동적 날짜 계산
        _today = datetime.date.today()
        _current_year = _today.year
        _current_quarter = f"{_current_year}-Q{(_today.month - 1) // 3 + 1}"
        _prev_q_month = _today.month - 3
        _prev_q_year = _current_year
        if _prev_q_month <= 0:
            _prev_q_month += 12
            _prev_q_year -= 1
        _prev_quarter = f"{_prev_q_year}-Q{(_prev_q_month - 1) // 3 + 1}"

        # 시장 데이터 컨텍스트 (payload에서 주입)
        market_context = payload.get("market_context", {}) or {}
        rates_txt = market_context.get("rates_summary", "")
        levelten_txt = market_context.get("levelten_summary", "")
        # BESS 소스 priority: LevelTen Storage (official) 1순위, AI Research (fallback) 2순위
        levelten_storage_txt = market_context.get("levelten_storage_summary", "")  # 공식 ISO-level
        bess_tolling_txt = market_context.get("bess_tolling_summary", "")          # AI Research duration별
        our_bess_duration = market_context.get("our_bess_duration", 4)
        # LevelTen 커버 여부 + 지역 해석 (WECC sub-region, SERC 등)
        lt_covered = market_context.get("levelten_covered", True)  # 기본 True (기존 프로젝트 호환)
        region_display = market_context.get("region_display", "")  # "WECC Rocky Mountain (UT)" 등
        sub_region = market_context.get("sub_region", "")          # WECC_RM, WECC_DSW, etc.
        continental_avg_txt = market_context.get("continental_avg_summary", "")  # Market-Averaged Continental (대용 비교용)

        market_block = ""
        if rates_txt or levelten_txt or levelten_storage_txt or bess_tolling_txt or continental_avg_txt:
            market_block = "=== CURRENT MARKET DATA (most recent; use this INSTEAD of training knowledge) ===\n"
            if rates_txt:
                market_block += f"  Interest Rates: {rates_txt}\n"
            # 지역 해석 명시
            if region_display:
                market_block += f"  Project Region: {region_display}"
                if lt_covered:
                    market_block += " [LevelTen 직접 커버 ISO]\n"
                else:
                    market_block += f" [LevelTen 미커버 — 대용 비교 필요]\n"
            if levelten_txt:
                market_block += f"  LevelTen PPA Benchmark (Solar): {levelten_txt}\n"
                if lt_covered:
                    market_block += "  → Our ISO가 LevelTen에 있음. USE IT to compare against project PPA directly.\n"
                    market_block += "    Reference: 'PPA $X.XX vs LevelTen P25 $Y.YY in {ISO}'.\n"
                else:
                    market_block += "  → Our region is NOT in LevelTen. Use as market context reference only.\n"
            # 대용 비교 (WECC/SERC 등 LevelTen 미커버 지역)
            if not lt_covered and continental_avg_txt:
                market_block += f"  Market-Averaged Continental Index (대용 비교용): {continental_avg_txt}\n"
                market_block += f"  → Use this as PRIMARY benchmark since {sub_region or 'project region'} has NO direct LevelTen coverage.\n"
                market_block += "  → Cite as 'LevelTen Market-Averaged Continental (전 대륙 ISO 평균, 대용치)'.\n"
                market_block += "  → Explicitly note '해당 지역 공식 P25 데이터 없음 → 대륙 평균 대비 비교' in risk commentary.\n"
            # Priority 1: LevelTen Storage (official, ISO-level)
            if levelten_storage_txt:
                market_block += f"  LevelTen Storage Index (OFFICIAL tolling offers, Q4 2025): {levelten_storage_txt}\n"
                market_block += f"  → Project BESS duration: {our_bess_duration}h (ISO-level price applies broadly, consider duration fit)\n"
                market_block += "  → USE THIS OFFICIAL DATA as primary BESS benchmark. Cite as 'LevelTen 공식 Storage Index'.\n"
                market_block += "  → If project toll EXCEEDS ISO P75 → risk 'BESS Toll 시장 상단 초과' (severity: Critical if >20% over, Watch if slight).\n"
                market_block += "  → If project toll is BELOW ISO P25 → positive flag '보수적 산정'.\n"
            # Priority 2: AI Research (fallback for non-LevelTen ISOs: ISO-NE/NYISO/WECC_*/SERC, or duration-level detail)
            if bess_tolling_txt:
                if levelten_storage_txt:
                    market_block += f"  AI Research Duration Detail (supplementary — LevelTen only provides ISO-level): {bess_tolling_txt}\n"
                    market_block += f"  → Use ONLY to add duration-specific nuance ({our_bess_duration}h). LevelTen ISO-level is primary.\n"
                    market_block += "  → Caveat: 'duration 세부는 AI 추정치'.\n"
                else:
                    # LevelTen 없는 지역 (WECC_*, ISO-NE, NYISO, SERC)
                    market_block += f"  BESS Tolling Estimate (AI Research — {sub_region or 'non-LevelTen region'}): {bess_tolling_txt}\n"
                    market_block += f"  → Project BESS duration: {our_bess_duration}h\n"
                    market_block += "  → CAVEAT: '시장 추정치, 공식 index 아님' when citing.\n"
                    if sub_region and sub_region.startswith("WECC"):
                        market_block += f"  → For {sub_region}: reference relevant utility RFPs (PacifiCorp IRP, URC, APS, Xcel Colorado, etc.) if AI Research provided commentary.\n"
            market_block += "\n"

        # 경제성 지표 추출 (Unlevered vs WACC 비교용)
        unlev_irr = current_metrics.get("unlevered_irr_pct")
        wacc_val  = current_metrics.get("wacc_pct")
        wacc_block = ""
        if unlev_irr is not None and wacc_val is not None:
            wacc_block = (
                f"  Unlevered Pre-Tax IRR : {unlev_irr}% (project-level)\n"
                f"  WACC                  : {wacc_val}% (hurdle)\n"
                f"  Value Creation        : Unlev - WACC = "
                f"{'POSITIVE' if float(unlev_irr)>float(wacc_val) else 'NEGATIVE'}\n"
            )

        prompt = (
        "You are the head of Investment Committee at Hanwha Energy USA (HEUH), "
        "a renewable energy developer whose sole business model is: develop → sell at NTP (before COD). "
        "The IC decision: should we continue spending development capital on this project?\n\n"

        f"TODAY'S DATE: {_today.isoformat()} (current quarter: {_current_quarter}, prior: {_prev_quarter}).\n\n"

        "═══ KNOWN REGULATORY & OPERATIONAL FACTS (treat as GIVEN; do not second-guess) ═══\n"
        "1. ITC Section 48E — Solar PV:\n"
        "   - 'Beginning of Construction' (BOC) is a LEGAL construct, not physical construction start.\n"
        "   - BOC deadline: July 4, 2026 — established via Physical Work Test (on-site or off-site binding work).\n"
        "   - Continuity Safe Harbor preserved: if BOC is established, project has until Dec 31, 2030 (4 years) to reach PIS.\n"
        "   - Projects missing BOC by July 4, 2026 must be Placed-in-Service by Dec 31, 2027.\n"
        "2. ITC Section 48E — BESS (SEPARATE TRACK from PV):\n"
        "   - Begin Construction by Dec 31, 2033 → 100% ITC\n"
        "   - 2034 → 75%, 2035 → 50%, 2036 → expires\n"
        "   - BESS is NOT subject to the 2026 solar cliff. Do NOT flag BESS ITC as imminent risk.\n"
        "3. HEUH Business Model & BOC Status:\n"
        "   - HEUH develops → sells at NTP (pre-COD). Post-COD execution risk does NOT affect IC decision.\n"
        "   - HEUH has established BOC for its project pool via Physical Work Test, managed by its compliance team.\n"
        "   - Individual project matching to BOC pool is operational matter — do NOT flag as financial risk.\n"
        "   - Post-BOC physical construction schedule is flexible within 4-year Continuity Safe Harbor.\n"
        "4. FEOC (Foreign Entity of Concern): compliance checklist item — do NOT use as verdict driver.\n\n"

        f"PROJECT: {proj_name} | Size: {pv_mwac} MWac\n"
        f"FINANCIAL SUMMARY: {context}\n"
        f"PROJECT METADATA: {proj_ctx}\n"
        f"ANNUAL SPONSOR CF (Y1-Y10): {cf_text}\n\n"

        + market_block +

        "=== INVESTMENT THRESHOLDS (firm hurdles) ===\n"
        f"  Primary   · Sponsor IRR ≥ {irr_thr}% (After-TE-Flip, Full Life) — 매수자 요구 수익률\n"
        f"  Secondary · Dev Margin  ≥ {margin_thr} c/Wp — HEUH 내부 마진 기준\n"
        "  Both must PASS for IC approval.\n\n"

        "=== CURRENT PROJECT METRICS ===\n"
        f"  Sponsor IRR : {curr_irr}% ({curr_irr_basis})\n"
        f"  Dev Margin  : {curr_margin} c/Wp\n"
        + (f"  Sponsor NPV : ${sponsor_npv_m}M (discounted at {irr_thr}% hurdle)\n" if sponsor_npv_m is not None else "")
        + (f"  Project NPV : ${project_npv_m}M (discounted at WACC)\n" if project_npv_m is not None else "")
        + wacc_block +
        f"  ITC Rate    : {curr_itc}%\n"
        f"  PPA Term    : {ppa_term} yrs | Toll Term: {toll_term} yrs\n\n"

        "═══ VERDICT FRAMEWORK (PURE ECONOMICS ONLY) ═══\n"
        "The verdict is determined ONLY by economic criteria. Development risks are monitoring items and do NOT affect verdict.\n\n"
        "Economic criteria:\n"
        "  1. Dev Margin vs threshold (primary: HEUH's exit value)\n"
        "  2. Sponsor IRR (After-TE-Flip, Full Life) vs threshold (market-clearing for buyer)\n"
        "  3. Unlevered IRR vs WACC (true value creation — leverage-independent)\n\n"
        "VERDICT RULES:\n"
        "  PROCEED:\n"
        "    - Dev Margin ≥ threshold AND Sponsor IRR ≥ threshold AND Unlev IRR > WACC\n"
        "    - Express threshold headroom explicitly if positive (e.g., '+1.5%p 여유')\n"
        "  RECUT:\n"
        "    - 1~2 criteria near miss (gap < 1.5%p from threshold) AND recoverable via negotiation\n"
        "    - Typical levers: PPA price revision, CAPEX reduction, TE/debt terms\n"
        "  STOP:\n"
        "    - Multiple criteria missed OR Unlev IRR < WACC (value destruction)\n"
        "    - Unrecoverable: gap too wide to close via normal levers\n\n"

        "═══ RISK ANALYSIS (monitoring only — NOT verdict driver) ═══\n"
        "Identify project-specific risks AI can assess:\n"
        "  - EPC price adequacy: $/Wdc vs current market (use supplied MARKET DATA if provided)\n"
        "  - ISO / interconnection queue risk based on ISO and state\n"
        "  - PPA market competitiveness vs supplied LevelTen P25 data (if given)\n"
        "  - Construction timeline vs PIS deadlines (Solar PV: 4-year continuity → Dec 31, 2030 PIS if BOC established; BESS: flexible to 2033)\n"
        "  - BESS replacement CAPEX / augmentation assumption sanity\n"
        "DO NOT generate risks for:\n"
        "  - Safe Harbor matching or BOC status (handled separately as fixed checklist item)\n"
        "  - FEOC compliance (handled separately as fixed checklist item)\n"
        "  - BESS ITC expiry (not imminent — 2033+ horizon)\n"
        "  - 'Must begin physical construction by 2026' — INCORRECT; BOC is a legal construct already managed via HEUH's Physical Work Test completion\n"
        "  - Generic 'market uncertainty' or 'policy risk' without specifics\n\n"

        "═══ LANGUAGE ═══\n"
        + ("ALL text fields in KOREAN (한국어).\n"
           "\n"
           "CRITICAL — Korean ENDING STYLE (IC memo convention, formal & concise):\n"
           "Use short nominal/verbal endings, NOT 존대체 (하다/한다) nor 합쇼체 (합니다).\n"
           "Required endings:\n"
           "  - 명사형 종결: '~ 충족', '~ 확인', '~ 권고', '~ 필요', '~ 가능', '~ 부족'\n"
           "  - 축약 서술: '~됨', '~함', '~임', '~없음', '~확보됨'\n"
           "Examples (GOOD):\n"
           "  ✓ '개발 마진 20.0 c/Wp로 기준 대비 +10.0%p 여유 확보'\n"
           "  ✓ '가중평균자본비용 대비 +0.88%p 상회로 가치 창출 확인'\n"
           "  ✓ 'PPA 재협상 또는 CAPEX 3% 절감 필요'\n"
           "  ✓ '경제성 기준 3개 모두 충족, 개발자본 투입 계속 권고'\n"
           "Examples (BAD — do NOT use):\n"
           "  ✗ '~제공한다' (→ '~제공')\n"
           "  ✗ '~확인된다' (→ '~확인됨')\n"
           "  ✗ '~충족한다' (→ '~충족')\n"
           "  ✗ '~하도록 한다' (→ '~권고')\n"
           "  ✗ '~해야 합니다' / '~할 수 있습니다' (too formal/verbose)\n"
           "Maintain consistency — ALL sentences end in the nominal/concise style.\n"
           "\n"
           "CRITICAL — INDUSTRY TERMINOLOGY:\n"
           "  ✗ '증설' (WRONG — means 'capacity expansion')\n"
           "  ✓ 'Augmentation' (English preferred, industry standard)\n"
           "  ✓ '배터리 교체 (용량 유지)' or '배터리 보강 (성능 유지)' (if Korean needed)\n"
           "  Augmentation = replacing/adding cells to MAINTAIN capacity over degradation,\n"
           "  NOT adding new capacity. Never call it '증설'.\n"
           "\n"
           "Only 'verdict' (PROCEED/RECUT/STOP) and 'verdict_color' (green/amber/red) stay English.\n"
           "Financial numbers with units can stay English-style (e.g., '10.38%', '$68.82/MWh').\n"
           "DO NOT mix languages within a single field.\n"
           if payload.get("lang","en")=="kr" else
           "ALL text fields in ENGLISH only. Formal institutional investor tone.\n")
        + "\n"
        "Be direct. Cite specific numbers. No hedging.\n\n"

        "Respond ONLY with valid JSON (no markdown, no code blocks).\n"
        "Required keys:\n"
        "  verdict: \"PROCEED\" | \"RECUT\" | \"STOP\"\n"
        "  verdict_color: \"green\" | \"amber\" | \"red\"\n"
        "  threshold_status: {\n"
        "    margin_ok: bool, margin_gap: str,\n"
        "    irr_ok: bool, irr_gap: str,\n"
        "    wacc_spread_ok: bool, wacc_spread: str  (e.g., '+0.88%p' or '-1.20%p')\n"
        "  }\n"
        "  metrics: ONE compact line, under 120 chars, pipe-delimited.\n"
        "    Example: '199 MWac | 10.4% IRR | $39.8M Margin | $68.8 PPA | $836M CAPEX | 30% ITC'\n"
        "  sensitivity_en: dev margin upside/downside in English with c/Wp numbers\n"
        "  sensitivity_kr: same in Korean (nominal ending style)\n"
        "  thesis: 3-4 sentence economic rationale (경제성 수치 기반 근거)\n"
        "  risks: array of {title, severity: Critical|Watch|OK, detail}\n"
        "    (project-specific only; Safe Harbor/FEOC/BESS ITC are handled separately)\n"
        "  rec: 2-3 sentence actionable recommendation (경제성 관점)\n"
        "All strings double-quoted. No trailing commas. No extra text outside JSON.\n"
        "NOTE: Do NOT include 'dev_ic' field — it has been removed from the schema."
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 4000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=45
    )
    if resp.status_code != 200:
        raise HTTPException(500, f"Claude API 오류: {resp.text[:200]}")

    data = resp.json()
    text = "".join(b.get("text","") for b in data.get("content",[]))

    # JSON 정제 — 코드블록, 줄바꿈, 특수문자 처리
    import re as _re
    clean = text.strip()
    clean = _re.sub(r"```(?:json)?\s*", "", clean).strip()
    clean = clean.strip("`")
    # { ... } 범위만 추출
    start = clean.find("{")
    end   = clean.rfind("}") + 1
    if start >= 0 and end > start:
        clean = clean[start:end]

    # ── 고정 규정 준수 체크리스트 2개 항목을 응답에 주입 ───────
    # AI가 판단하는 risks와 완전히 분리된, 모든 프로젝트 공통 체크리스트
    is_kr = payload.get("lang", "en") == "kr"

    if is_kr:
        compliance_checklist = [
            {
                "title": "ITC BOC(Beginning of Construction) 매칭 확인",
                "severity": "Watch",
                "detail": (
                    "HEUH는 Physical Work Test 방식으로 BOC 요건을 확보하여 관리 중. "
                    "본 프로젝트가 기확보된 BOC pool과 매칭되는지 NTP 전 확인 권고. "
                    "매칭 확보 시 Continuity Safe Harbor에 따라 2030년 말까지 PIS 여유."
                )
            },
            {
                "title": "FEOC 공급망 적격성 검토",
                "severity": "Watch",
                "detail": (
                    "OBBBA에 따라 2026년 착공 프로젝트는 비PFE(중국/러시아/이란/북한 외) "
                    "부품 비중 요건 적용: PV ≥40%, BESS ≥55% (매년 5%p 상향). "
                    "EPC 계약 체결 전 배터리 셀·PV 모듈 원산지 증빙 확보 필요."
                )
            }
        ]
    else:
        compliance_checklist = [
            {
                "title": "ITC BOC Matching Verification",
                "severity": "Watch",
                "detail": (
                    "HEUH has established BOC (Beginning of Construction) for its project pool "
                    "via Physical Work Test, managed by its compliance team. Verify this project "
                    "is matched to the secured BOC pool before NTP. Once matched, Continuity Safe "
                    "Harbor extends PIS to Dec 31, 2030."
                )
            },
            {
                "title": "FEOC Supply Chain Compliance Review",
                "severity": "Watch",
                "detail": (
                    "Under OBBBA, 2026-start projects face non-PFE (China/Russia/Iran/DPRK excluded) "
                    "content thresholds: PV ≥40%, BESS ≥55% (ramping +5%p annually). Verify battery "
                    "cell and PV module country-of-origin documentation before EPC contract."
                )
            }
        ]

    # JSON 파싱해서 risks 배열 앞에 삽입
    try:
        import json as _json
        parsed = _json.loads(clean)
        ai_risks = parsed.get("risks", []) or []
        # 컴플라이언스 체크리스트를 최상단, AI 리스크를 그 뒤에
        parsed["risks"] = compliance_checklist + ai_risks
        # 구분 위해 플래그 추가 (프론트엔드에서 활용 가능)
        parsed["compliance_count"] = len(compliance_checklist)
        clean = _json.dumps(parsed, ensure_ascii=False)
    except Exception as _e:
        # 파싱 실패 시 원본 그대로 반환 (프론트가 처리)
        print(f"[analyze-cf] JSON merge failed: {_e}", flush=True)

    return {"ok": True, "result": clean}


@app.post("/valuation/{project_id}/save")
async def save_valuation_version(
    project_id: str,
    payload: dict,
    user=Depends(get_current_user)
):
    """버전 저장 → 즉시 저장 (승인 flow 제거). 100개 한도."""
    safe_id = project_id.replace("/", "_").replace(".", "_")
    ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    payload["uploaded_by"] = user["email"]
    payload["uploaded_at"] = datetime.datetime.now().isoformat()
    # 승인 flow 제거 → 즉시 "saved" 상태
    payload["status"] = "saved"
    payload["requested_by"] = user["email"]
    # "approver" 필드 레거시 호환: 존재하면 "shared_with"로 마이그레이트
    if "approver" in payload and payload["approver"] and "shared_with" not in payload:
        payload["shared_with"] = payload["approver"]

    fb_put(f"valuation/{safe_id}/versions/{ts}", payload)
    fb_put(f"valuation/{safe_id}/latest", payload)

    # 100개 한도 — 초과 시 가장 오래된 것 삭제
    versions = fb_read(f"valuation/{safe_id}/versions") or {}
    keys = sorted(versions.keys())
    if len(keys) > 100:
        for old_key in keys[:len(keys)-100]:
            try:
                requests.delete(
                    f"{FB_URL}/valuation/{safe_id}/versions/{old_key}.json",
                    params=fb_auth_param(),
                    timeout=5
                )
            except Exception:
                pass

    return {"ok": True, "timestamp": ts}


# 레거시 승인/반려 엔드포인트 — 하위호환 유지하되 no-op화 (존재하는 pending 버전 정리용)
@app.post("/valuation/{project_id}/versions/{ts}/approve")
def approve_version(project_id: str, ts: str, user=Depends(require_admin)):
    """[Deprecated] 승인 flow 제거됨. 하위호환용: pending을 saved로 마이그레이트."""
    safe_id = project_id.replace("/", "_").replace(".", "_")
    fb_patch(f"valuation/{safe_id}/versions/{ts}", {
        "status": "saved",
        "approved_by": user["email"],
        "approved_at": datetime.datetime.now().isoformat()
    })
    return {"ok": True}


@app.post("/valuation/{project_id}/versions/{ts}/reject")
def reject_version(project_id: str, ts: str, body: dict = {}, user=Depends(require_admin)):
    """[Deprecated] 승인 flow 제거됨. 하위호환용: 버전 삭제."""
    safe_id = project_id.replace("/", "_").replace(".", "_")
    fb_patch(f"valuation/{safe_id}/versions/{ts}", {
        "status": "rejected",
        "rejected_by": user["email"],
        "rejected_at": datetime.datetime.now().isoformat(),
        "reject_reason": body.get("reason", "")
    })
    return {"ok": True}
